from collections import defaultdict
from typing import Sequence, Callable
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analytics.excel_executive import C_CAT_BG, C_CAT_FG, C_SUBCAT_BG, C_SUBCAT_FG, TAB_COLORS

# Colores y formatos
C_ROJO_BG = "DC2626"
C_ESTANCADO_BG = "2563EB"
C_ROTACION_BG = "059669"
C_DEPT_BG = "E5E7EB"
C_NOTE_FG = "6B7280"

INT_FMT = "#,##0"
DEC1_FMT = "0.0"
MONEY_FMT = '"S/" #,##0.00'
PCT_FMT = "0.0%"

FILL_ROJO = PatternFill("solid", fgColor="FEE2E2") # red-100
FILL_NARANJA = PatternFill("solid", fgColor="FFEDD5") # orange-100
FILL_AMARILLO = PatternFill("solid", fgColor="FEF9C3") # yellow-100
FILL_VERDE = PatternFill("solid", fgColor="D1FAE5") # green-100
FILL_AZUL = PatternFill("solid", fgColor="DBEAFE") # blue-100
FILL_GRIS = PatternFill("solid", fgColor="F3F4F6") # gray-100

_BORDER_ROW = Border(bottom=Side(style="thin", color="E5E7EB"), top=Side(style="thin", color="E5E7EB"))



def _banner(ws: Worksheet, titulo: str, *, brand_name: str, sucursal: str | None, ncols: int, color_bg: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c1 = ws.cell(row=1, column=1, value=f"{brand_name} - Análisis Gerencial")
    c1.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor="1F2937")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    suc = sucursal if sucursal else "Todas las sucursales"
    c2 = ws.cell(row=2, column=1, value=f"{titulo} | {suc}")
    c2.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor=color_bg)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 10

def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Evita nombres duplicados o muy largos
    base = name[:28]
    if base not in used:
        used.add(base)
        return base
    for i in range(1, 100):
        cand = f"{base[:25]} ({i})"
        if cand not in used:
            used.add(cand)
            return cand
    return base

def _num(v) -> float:
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

def _safe_date(d) -> str:
    return str(d) if d else ""

def _tabla_jerarquica(
    ws: Worksheet,
    start_row: int,
    headers: Sequence[str],
    fmts: Sequence[str],
    items: list[dict],
    cols_to_sum: list[int],
    row_builder: Callable[[dict], tuple],
    row_fill: Callable[[dict], PatternFill | None] | None = None,
    bold_cols: set[int] = frozenset(),
) -> int:
    cat_groups = defaultdict(lambda: defaultdict(list))
    for item in items:
        cat = item.get("Categoría", "(sin categoría)") or "(sin categoría)"
        sub = item.get("Subcategoría", "(sin subcategoría)") or "(sin subcategoría)"
        cat_groups[cat][sub].append(item)
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_CAT_FG)
        cell.fill = PatternFill("solid", fgColor=C_CAT_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=C_DEPT_BG))
    ws.row_dimensions[start_row].height = 24
    
    r = start_row + 1
    
    for cat_name, sub_groups in sorted(cat_groups.items()):
        cat_totals = {c: 0.0 for c in cols_to_sum}
        for sub_name, prods in sub_groups.items():
            for p in prods:
                row_tup = row_builder(p)
                for c in cols_to_sum:
                    val = row_tup[c - 6]
                    if isinstance(val, (int, float)):
                        cat_totals[c] += val
        
        cat_fill = PatternFill("solid", fgColor=C_CAT_BG)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = cat_fill
            cell.border = _BORDER_ROW
        
        c = ws.cell(row=r, column=1, value="📂 Categoría")
        c.font = Font(name="Calibri", size=11, bold=True, color=C_CAT_FG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        
        c = ws.cell(row=r, column=2, value=cat_name)
        c.font = Font(name="Calibri", size=11, bold=True, color=C_CAT_FG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        
        for c_idx in cols_to_sum:
            c = ws.cell(row=r, column=c_idx, value=cat_totals[c_idx])
            c.font = Font(name="Calibri", size=11, bold=True, color=C_CAT_FG)
            fmt = fmts[c_idx - 1]
            if fmt == "int":
                c.number_format = INT_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "dec1":
                c.number_format = DEC1_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "money":
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "pct":
                c.number_format = PCT_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.row_dimensions[r].height = 22
        ws.row_dimensions[r].outline_level = 1
        r += 1
        
        for sub_name, prods in sorted(sub_groups.items()):
            sub_totals = {c: 0.0 for c in cols_to_sum}
            for p in prods:
                row_tup = row_builder(p)
                for c in cols_to_sum:
                    val = row_tup[c - 6]
                    if isinstance(val, (int, float)):
                        sub_totals[c] += val
            
            sub_fill = PatternFill("solid", fgColor=C_SUBCAT_BG)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.fill = sub_fill
                cell.border = _BORDER_ROW
            
            c = ws.cell(row=r, column=1, value="📁 Subcategoría")
            c.font = Font(name="Calibri", size=10, bold=True, color=C_SUBCAT_FG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            
            c = ws.cell(row=r, column=2, value=cat_name)
            c.font = Font(name="Calibri", size=10, bold=True, color=C_SUBCAT_FG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            
            c = ws.cell(row=r, column=3, value=sub_name)
            c.font = Font(name="Calibri", size=10, bold=True, color=C_SUBCAT_FG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            
            for c_idx in cols_to_sum:
                c = ws.cell(row=r, column=c_idx, value=sub_totals[c_idx])
                c.font = Font(name="Calibri", size=10, bold=True, color=C_SUBCAT_FG)
                fmt = fmts[c_idx - 1]
                if fmt == "int":
                    c.number_format = INT_FMT
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif fmt == "dec1":
                    c.number_format = DEC1_FMT
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif fmt == "money":
                    c.number_format = MONEY_FMT
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif fmt == "pct":
                    c.number_format = PCT_FMT
                    c.alignment = Alignment(horizontal="right", vertical="center")
            
            ws.row_dimensions[r].height = 22
            ws.row_dimensions[r].outline_level = 2
            r += 1
            
            for i, p in enumerate(prods):
                fill = row_fill(p) if row_fill else None
                row_tup = row_builder(p)
                
                base_tup = (
                    "▸ Producto",
                    cat_name,
                    sub_name,
                    p.get("Código SKU", ""),
                    p.get("Producto", "")
                )
                full_tup = base_tup + row_tup
                
                for c_idx, val in enumerate(full_tup, 1):
                    cell = ws.cell(row=r, column=c_idx, value=val)
                    cell.font = Font(name="Calibri", size=10, bold=(c_idx - 1) in bold_cols, color="1F2937")
                    if c_idx == 4:
                        cell.font = Font(name="Consolas", size=9, color="4B5563")
                    
                    cell.border = _BORDER_ROW
                    if fill is not None:
                        cell.fill = fill
                    elif i % 2 == 1:
                        cell.fill = PatternFill("solid", fgColor="FAFAFA")
                    
                    fmt = fmts[c_idx - 1]
                    if fmt == "int":
                        cell.number_format = INT_FMT
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif fmt == "dec1":
                        cell.number_format = DEC1_FMT
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif fmt == "money":
                        cell.number_format = MONEY_FMT
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif fmt == "pct":
                        cell.number_format = PCT_FMT
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                ws.row_dimensions[r].outline_level = 3
                r += 1

    for c_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 15
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    
    return r

def _add_dept_sheets_jerarquico(
    wb: Workbook,
    grupos: list[tuple[str, list]],
    *,
    brand_name: str,
    sucursal: str | None,
    color_bg: str,
    icono: str,
    headers: Sequence[str],
    fmts: Sequence[str],
    cols_to_sum: list[int],
    row_builder: Callable[[dict], tuple],
    row_fill: Callable[[dict], PatternFill | None] | None = None,
    bold_cols: set[int] = frozenset(),
    nota_de: Callable[[list], str] | None = None,
) -> None:
    used: set[str] = set()
    ncols = len(headers)
    for rank, (dept, its) in enumerate(grupos, 1):
        ws = wb.create_sheet(_safe_sheet_name(f"{rank}. {dept}", used))
        ws.sheet_properties.tabColor = TAB_COLORS[(rank - 1) % len(TAB_COLORS)]
        _banner(ws, f"{icono}  {dept}", brand_name=brand_name,
                sucursal=sucursal, ncols=ncols, color_bg=color_bg)
        start = 5
        if nota_de:
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=ncols)
            c = ws.cell(row=5, column=1, value=nota_de(its))
            c.font = Font(name="Calibri", size=10, italic=True, color=C_NOTE_FG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[5].height = 20
            start = 7
        
        _tabla_jerarquica(
            ws, start, headers, fmts, its,
            cols_to_sum=cols_to_sum,
            row_builder=row_builder,
            row_fill=row_fill,
            bold_cols=bold_cols
        )

# ══════════════════════════════════════════════════════════════════════════
# 1) POR AGOTARSE
# ══════════════════════════════════════════════════════════════════════════
def build_por_agotarse_workbook(
    rows: list[dict], dias_alerta: int, sucursal: str | None, brand_name: str,
    cobertura_objetivo_dias: int = 30,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    items = []
    for r in rows:
        unds90 = _num(r.get("Unds Vend (90d)"))
        if unds90 <= 0:
            continue
        stock = _num(r.get("Stock Disp"))
        ritmo = unds90 / 90.0
        dias_agota = int(stock / ritmo) if ritmo > 0 else 999
        if dias_agota <= dias_alerta:
            r["_dias_agota"] = dias_agota
            r["_ritmo"] = ritmo
            # Misma fórmula que el dashboard de compras (/compras-catalogo):
            # reponer hasta cobertura_objetivo_dias usando velocidad reciente 30d
            # (fallback: velocidad del lote).
            vel_30d = _num(r.get("Vel últimos 30d"))
            vel_90d = _num(r.get("Velocidad (uds/día)"))
            vel_ref = vel_30d if vel_30d > 0 else vel_90d
            r["_comprar"] = max(0, round(vel_ref * cobertura_objetivo_dias) - int(stock))
            items.append(r)
            
    dept_groups = defaultdict(list)
    for c in items:
        dept = str(c.get("Departamento") or "(sin departamento)")
        dept_groups[dept].append(c)
        
    grupos = sorted(dept_groups.items(), key=lambda kv: len(kv[1]), reverse=True)
    
    headers = ["Tipo", "Categoría", "Subcategoría", "SKU", "Producto",
               "Stock actual", "Vende por día", "Se agota en (días)",
               "Vendido 90 días", f"Comprar sugerido ({cobertura_objetivo_dias} días)",
               "Último Ingreso", "Llegó hace (días)"]
    fmts = ["text", "text", "text", "text", "text",
            "int", "dec1", "int", "int", "int", "text", "int"]

    _add_dept_sheets_jerarquico(
        wb, grupos, brand_name=brand_name, sucursal=sucursal,
        color_bg=C_ROJO_BG, icono="🚨", headers=headers, fmts=fmts,
        cols_to_sum=[6, 9, 10],
        row_builder=lambda c: (
            _num(c.get("Stock Disp")),
            c.get("_ritmo", 0),
            c.get("_dias_agota", 0),
            _num(c.get("Unds Vend (90d)")),
            c.get("_comprar", 0),
            _safe_date(c.get("Últ. Recepción")),
            int(_num(c.get("Llegó hace (días)"))) if c.get("Llegó hace (días)") else None,
        ),
        row_fill=None, bold_cols={3, 5, 7},
        nota_de=lambda its: (
            f"Mostrando {len(its)} productos con stock bajo. "
            f"'Comprar sugerido' = cantidad para cubrir {cobertura_objetivo_dias} días de venta "
            f"según velocidad reciente (igual al dashboard de compras)."
        )
    )
    if not wb.sheetnames:
        ws = wb.create_sheet("Sin datos")
        ws.cell(1, 1, value="No hay datos para mostrar.")
    return wb

# ══════════════════════════════════════════════════════════════════════════
# 2) ESTANCADOS
# ══════════════════════════════════════════════════════════════════════════
def build_estancados_workbook(
    rows: list[dict], costos: dict, dias_estancado: int, sucursal: str | None, brand_name: str
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    
    items = []
    for r in rows:
        dias_sin_vender = int(_num(r.get("Días sin Vender")))
        if dias_sin_vender >= dias_estancado:
            stock = _num(r.get("Stock Disp"))
            if stock > 0:
                costo = costos.get(str(r.get("Código SKU")), 0.0)
                r["_dias_sin_vender"] = dias_sin_vender
                r["_costo"] = costo
                r["_inmovilizado"] = stock * costo
                r["_que_hacer"] = "LIQUIDAR" if dias_sin_vender >= 180 else ("OFERTAR" if dias_sin_vender >= 90 else "OBSERVAR")
                items.append(r)

    dept_groups = defaultdict(list)
    for c in items:
        dept = str(c.get("Departamento") or "(sin departamento)")
        dept_groups[dept].append(c)
        
    grupos = sorted(dept_groups.items(), key=lambda kv: sum(i["_inmovilizado"] for i in kv[1]), reverse=True)

    headers = ["Tipo", "Categoría", "Subcategoría", "SKU", "Producto",
               "Stock actual", "Días sin vender", "S/ inmovilizado", "Qué hacer",
               "Último Ingreso", "Llegó hace (días)"]
    fmts = ["text", "text", "text", "text", "text",
            "int", "int", "money", "text", "text", "int"]

    def _fill_est(it: dict) -> PatternFill | None:
        if it.get("_que_hacer", "") == "LIQUIDAR":
            return FILL_ROJO
        elif it.get("_que_hacer", "") == "OFERTAR":
            return FILL_NARANJA
        return None

    _add_dept_sheets_jerarquico(
        wb, grupos, brand_name=brand_name, sucursal=sucursal,
        color_bg=C_ESTANCADO_BG, icono="🧊", headers=headers, fmts=fmts,
        cols_to_sum=[6, 8],
        row_builder=lambda it: (
            _num(it.get("Stock Disp")),
            it.get("_dias_sin_vender", 0),
            it.get("_inmovilizado", 0.0),
            it.get("_que_hacer", ""),
            _safe_date(it.get("Últ. Recepción")),
            int(_num(it.get("Llegó hace (días)"))) if it.get("Llegó hace (días)") else None,
        ),
        row_fill=_fill_est, bold_cols={3, 5, 7},
        nota_de=lambda its: f"Hay {len(its)} productos estancados en este departamento (S/ {sum(it.get('_inmovilizado', 0) for it in its):,.2f})"
    )
    if not wb.sheetnames:
        ws = wb.create_sheet("Sin datos")
        ws.cell(1, 1, value="No hay datos para mostrar.")
    return wb

# ══════════════════════════════════════════════════════════════════════════
# 3) ROTACIÓN
# ══════════════════════════════════════════════════════════════════════════
def build_rotacion_workbook(
    rows: list[dict], sucursal: str | None, brand_name: str
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    
    items = []
    for r in rows:
        unds90 = _num(r.get("Unds Vend (90d)"))
        if unds90 <= 0 and _num(r.get("Stock Disp")) <= 0:
            continue
            
        if unds90 >= 30:
            rot_nivel = "🟢 RÁPIDA"
        elif unds90 >= 9:
            rot_nivel = "🔵 NORMAL"
        elif unds90 >= 2:
            rot_nivel = "🟡 LENTA"
        else:
            rot_nivel = "⚪ NO SE MUEVE"
            
        r["_rotacion"] = rot_nivel
        r["_ritmo"] = unds90 / 90.0
        r["_ventas_soles"] = _num(r.get("Vendido SKU S/"))
        items.append(r)
        
    dept_groups = defaultdict(list)
    for c in items:
        dept = str(c.get("Departamento") or "(sin departamento)")
        dept_groups[dept].append(c)
        
    grupos = sorted(dept_groups.items(), key=lambda kv: len(kv[1]), reverse=True)

    headers = ["Tipo", "Categoría", "Subcategoría", "SKU", "Producto",
               "Nivel de rotación", "Ritmo de venta",
               "Vend. 90d", "Ventas 90d S/", "Último Ingreso", "Llegó hace (días)"]
    fmts = ["text", "text", "text", "text", "text",
            "text", "text", "int", "money", "text", "int"]

    def _fill_rot(it: dict) -> PatternFill | None:
        rot = str(it.get("_rotacion", "")).upper()
        if "RÁPIDA" in rot:
            return FILL_VERDE
        elif "LENTA" in rot:
            return FILL_AMARILLO
        elif "NO SE MUEVE" in rot:
            return FILL_GRIS
        return None

    _add_dept_sheets_jerarquico(
        wb, grupos, brand_name=brand_name, sucursal=sucursal,
        color_bg=C_ROTACION_BG, icono="🔄", headers=headers, fmts=fmts,
        cols_to_sum=[8, 9],
        row_builder=lambda it: (
            it.get("_rotacion", ""),
            f"{it.get('_ritmo', 0):.1f} und/día",
            _num(it.get("Unds Vend (90d)")),
            it.get("_ventas_soles", 0.0),
            _safe_date(it.get("Últ. Recepción")),
            int(_num(it.get("Llegó hace (días)"))) if it.get("Llegó hace (días)") else None,
        ),
        row_fill=_fill_rot, bold_cols={3, 5, 8},
        nota_de=None
    )
    if not wb.sheetnames:
        ws = wb.create_sheet("Sin datos")
        ws.cell(1, 1, value="No hay datos para mostrar.")
    return wb
