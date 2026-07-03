"""
Builder ejecutivo (.xlsx) — formato Ventas & Catálogo / Reporte Diario.

Reemplazo de `excel_builder.build_workbook` para el módulo 04b. Inspirado en el
layout "COYA Analytics" pero re-paletado en naranja (Tailwind orange) para
branding HUDEC.

Layout:
    1) Resumen        — ranking 1..N por Departamento (Ventas, %, Tickets,
                        Ticket prom)
    2..N) Una hoja por Departamento, con cuerpo PLANO (sin outline):
          📂 Categoría → 📁 Subcategoría → ▸ Producto
          Filas estilizadas en niveles jerárquicos.

    Cada SubCat trae además:
      • Clasificación AGREGADA (la dominante entre sus SKUs)
      • Conteo "N ganadores · M en alerta"

    Productos dentro de SubCat marcada LIQUIDAR se resaltan como **rescatables**
    (banda dorada con borde izquierdo naranja).

Columnas de cada hoja de dept: ver DEPT_HEADERS (agrupadas por tema:
    Identificación → Ventas → Stock → Ciclo del lote → Diagnóstico).

Entrada (mismas que `build_workbook` para drop-in):
    cols  : list[str]    columnas que devuelve `service.run_matrix(...)`
    rows  : Iterable[tuple]  filas como tuplas en el mismo orden
    + meta (título, descripción, sucursal, etc.) para la portada del Resumen.

Salida: openpyxl.Workbook ya armado, listo para serializar.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ──────────────────────────────────────────────────────────────────────────
# Paleta NARANJA (Tailwind orange + grays). Equivalente al rosa/magenta COYA.
# ──────────────────────────────────────────────────────────────────────────
C_TITLE_BG    = "9A3412"  # orange-800   — Title big bar
C_TITLE_FG    = "FFFFFF"
C_META_BG     = "1F2937"  # gray-800     — meta subtítulo
C_META_FG     = "FDBA74"  # orange-300
C_TOTAL_BG    = "FFF7ED"  # orange-50
C_TOTAL_LBL   = "6B7280"  # gray-500
C_TOTAL_FG    = "9A3412"  # orange-800
C_NOTE_FG     = "6B7280"

C_HEADER_BG   = "1F2937"
C_HEADER_FG   = "FFFFFF"

C_DEPT_BG     = "C2410C"  # orange-700
C_DEPT_FG     = "FFFFFF"
C_DEPT_BORDER = "7C2D12"  # orange-900

C_CAT_BG      = "FED7AA"  # orange-200
C_CAT_FG      = "9A3412"  # orange-800

C_SUBCAT_BG   = "FFF7ED"  # orange-50
C_SUBCAT_FG   = "C2410C"  # orange-700

C_PROD_BG     = "FFFFFF"
C_PROD_BG_ALT = "FAFAFA"
C_PROD_FG     = "1F2937"
C_PROD_MONO   = "6B7280"

C_RANK_BG     = "FFF7ED"
C_RANK_FG     = "9A3412"

C_RESCATE_BG  = "FEF3C7"  # amber-100
C_RESCATE_FG  = "78350F"  # amber-900
C_RESCATE_BORDER = "F59E0B"  # amber-500

# Tones por clasificación (font, fill) — equivalentes a ToneBest/Good/...
TONE_BEST    = ("065F46", "A7F3D0")  # verde
TONE_GOOD    = ("155E75", "A5F3FC")  # cyan
TONE_WARN    = ("78350F", "FDE68A")  # amarillo
TONE_BAD     = ("7F1D1D", "FECACA")  # rojo
TONE_NEW     = ("365314", "D9F99D")  # lima
TONE_NEUTRAL = ("374151", "F3F4F6")  # gris

TAB_COLORS = [
    "9A3412", "C2410C", "EA580C", "F97316", "FB923C",
    "FDBA74", "FED7AA", "FFEDD5", "B45309", "78350F",
]


# ──────────────────────────────────────────────────────────────────────────
# Estilos pre-armados (instancias) — openpyxl exige instancias por celda.
# ──────────────────────────────────────────────────────────────────────────

def _font(color: str = "1F2937", size: int = 10, bold: bool = False,
          italic: bool = False, name: str = "Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_bottom(color: str = "000000", weight: str = "thin") -> Border:
    return Border(bottom=Side(style=weight, color=color))


def _border_left_thick(color: str = "F59E0B") -> Border:
    return Border(left=Side(style="medium", color=color))


# Formatos numéricos
NF_MONEY = '"S/ "#,##0.00'
NF_INT   = '#,##0'
NF_PCT   = '0.0%'
NF_DATE  = 'dd/mm/yyyy'


# ──────────────────────────────────────────────────────────────────────────
# Tone por clasificación (replica de classifyTone del frontend)
# ──────────────────────────────────────────────────────────────────────────

def classify_tone(label: str) -> tuple[str, str]:
    """(font_color, bg_color) según la clasificación textual del SKU."""
    L = (label or "").upper()
    if "ALTA ROTACIÓN" in L or "ALTA ROT" in L or "EXITOSO" in L or "JOYA" in L or "🏆" in label or "💎" in label:
        return TONE_BEST
    # ★ FIX 2026-06-10: tokens "REPONER *" precisos — "REPONER" a secas matcheaba
    #   "NO REPONER" / "EVALUAR ANTES DE REPONER" y pintaba de verde productos muertos.
    if any(k in L for k in ("REPONER YA", "REPONER POCO", "REPONER MENOS",
                            "AGOTADO CON DEMANDA", "POCO STOCK CON DEMANDA",
                            "MEDIA ROT", "POTENCIAL", "STOCK PREVIO", "ROTACIÓN ACTIVA",
                            "INVENTARIO SANO", "LOTE NUEVO VENDIENDO")):
        return TONE_GOOD
    if any(k in L for k in ("EXCESO", "AGOTAD", "BAJA ROT", "LENTA", "LENTO", "DESPERTÓ", "QUIEBRE", "RITMO PERDIDO")):
        return TONE_WARN
    if any(k in L for k in ("MUERTO", "LIQUIDAR", "DESCATAL", "FRACASO", "RESIDUO", "HISTÓRICO", "MARGINAL")):
        return TONE_BAD
    if any(k in L for k in ("NUEVO", "EMERGENTE", "RECIBIDO")):
        return TONE_NEW
    return TONE_NEUTRAL


def _is_winner(label: str) -> bool:
    """Un SKU es 'ganador' si su tone es Best o Good."""
    return classify_tone(label) in (TONE_BEST, TONE_GOOD)


def _is_alert(label: str) -> bool:
    """Un SKU es 'alerta' si su tone es Bad."""
    return classify_tone(label) == TONE_BAD


def _is_liquidate(label: str) -> bool:
    L = (label or "").upper()
    return any(k in L for k in ("MUERTO", "LIQUIDAR", "DESCATAL", "FRACASO", "RESIDUO", "MARGINAL", "HISTÓRICO"))


# ──────────────────────────────────────────────────────────────────────────
# Veredicto por SubCat/Cat: ¿Seguir comprando o descartar?
# Usa el mix de ventas entre SKUs comprables vs problemáticos.
# ──────────────────────────────────────────────────────────────────────────

BUCKET_COMPRABLE = (
    "ALTA ROTACIÓN", "ALTA ROT", "ROTACIÓN ACTIVA", "INVENTARIO SANO",
    "EXITOSO", "REABASTECIDO ACTIVO", "QUIEBRE STOCK", "POTENCIAL ACTIVO",
    # ★ FIX 2026-06-10: antes el token era "REPONER" a secas y matcheaba por
    #   substring "NO REPONER" (💤 DEMANDA EXTINTA, 🪦 LENTO CRÓNICO) y
    #   "EVALUAR ANTES DE REPONER" (📉 RITMO PERDIDO) → productos muertos
    #   contaban como SANOS en el veredicto de subcat/cat. Tokens precisos:
    "REPONER YA", "REPONER POCO", "REPONER MENOS",
    "AGOTADO CON DEMANDA", "POCO STOCK CON DEMANDA",
    "STOCK PREVIO", "MEDIA ROT", "LOTE AGOTADO RÁPIDO",
    # ★ FIX 2026-06-10: el SQL describe esta etiqueta como "sano" pero el rollup
    #   la contaba como neutro → subcats con su mejor SKU rotando bien salían
    #   🔴 REDUCIR/DESCARTAR (caso testigo: Toallitas Húmedas Básicas 29%→77%).
    "LOTE NUEVO VENDIENDO",
)
BUCKET_PROBLEMA = (
    "MUERTO", "RITMO PERDIDO", "EXCESO", "BAJA ROT", "DEMANDA EXTINTA",
    "MARGINAL", "HISTÓRICO", "FRACASO", "RESIDUO", "PÉRDIDA TOTAL",
    "LIQUIDAR", "DESCATAL",
    "LENTO CRÓNICO",  # ★ FIX 2026-06-10: "NO REPONER" — antes caía en comprable por substring
    "EXCESIVO",       # ★ FIX 2026-06-10: "🧊 STOCK EXCESIVO" no matchea "EXCESO" (EXCESIVO ≠ EXCESO)
)


def categorize_bucket(label: str) -> str:
    """3 buckets de decisión: comprable / problema / neutro."""
    L = (label or "").upper()
    for k in BUCKET_COMPRABLE:
        if k in L:
            return "comprable"
    for k in BUCKET_PROBLEMA:
        if k in L:
            return "problema"
    return "neutro"


def compute_veredicto(skus: dict) -> dict:
    """Calcula el veredicto de una SubCat/Cat según el mix de ventas.
    Devuelve {veredicto, razon, pct_sanas, n_estrellas, n_problemas_stock, tone}.

    Reglas:
      • ≥70% ventas en SKUs sanos → 🟢 SEGUIR COMPRANDO
      • 40-70% → 🟡 EVALUAR
      • <40% → 🔴 REDUCIR/DESCARTAR
    """
    ventas_total = sum(s["ventas_s"] for s in skus.values())
    ventas_comprable = sum(s["ventas_s"] for s in skus.values()
                           if categorize_bucket(s["clasif"]) == "comprable")
    n_estrellas = sum(1 for s in skus.values()
                      if categorize_bucket(s["clasif"]) == "comprable")
    n_problemas_stock = sum(1 for s in skus.values()
                            if categorize_bucket(s["clasif"]) == "problema"
                            and s["stock"] > 0)

    if ventas_total <= 0:
        return {"veredicto": "— sin ventas", "razon": "no hubo movimiento",
                "pct_sanas": 0, "n_estrellas": n_estrellas,
                "n_problemas_stock": n_problemas_stock, "tone": TONE_NEUTRAL}

    pct = (ventas_comprable / ventas_total) * 100
    if pct >= 70:
        return {"veredicto": "🟢 SEGUIR COMPRANDO",
                "razon": f"{pct:.0f}% de ventas en SKUs sanos",
                "pct_sanas": pct, "n_estrellas": n_estrellas,
                "n_problemas_stock": n_problemas_stock, "tone": TONE_BEST}
    if pct >= 40:
        return {"veredicto": "🟡 EVALUAR",
                "razon": f"solo {pct:.0f}% en SKUs sanos · revisar surtido",
                "pct_sanas": pct, "n_estrellas": n_estrellas,
                "n_problemas_stock": n_problemas_stock, "tone": TONE_WARN}
    return {"veredicto": "🔴 REDUCIR/DESCARTAR",
            "razon": f"solo {pct:.0f}% en SKUs sanos · capital estancado",
            "pct_sanas": pct, "n_estrellas": n_estrellas,
            "n_problemas_stock": n_problemas_stock, "tone": TONE_BAD}


def aggregate_cat_skus(cat_data: dict) -> dict:
    """Aplana todos los SKUs de todas las SubCats de una Cat para veredicto agregado."""
    out = {}
    for sub_skus in cat_data.values():
        out.update(sub_skus)
    return out


def _coverage_cell(stock: float, velocidad: float, cobertura_txt: str,
                   merged: bool) -> dict:
    """Calcula la celda 'Días Stock' (días de inventario que le quedan al SKU).

    Estrategia:
      • stock <= 0                → 'Agotado' (rojo).
      • 1 sola sucursal + texto   → reusar el texto del SQL ('110 días') tal
        cual lo muestra la matriz en pantalla (consistencia visual).
      • multi-sucursal consolidado→ recalcular round(stock_total / vel_total),
        porque el stock mostrado es la suma de todas las sucursales.
      • con stock pero sin venta  → 'Sin rotación' (gris).

    Devuelve: {value, is_number, font, fill} para que el caller lo pinte.
    Color por urgencia: ≤7d rojo, ≤30d ámbar, resto normal.
    """
    days: int | None = None
    label: str | None = None

    if stock <= 0:
        label = "Agotado"
    elif not merged and cobertura_txt:
        m = re.search(r"-?\d+", cobertura_txt)
        if m:
            days = int(m.group())
        else:
            label = cobertura_txt  # ej. 'Agotado'/'Sin rotación' que ya trae el SQL
    elif velocidad and velocidad > 0:
        days = max(1, round(stock / velocidad))
    else:
        label = "Sin rotación"

    # Estilo según urgencia
    if label == "Agotado" or (days is not None and days <= 7):
        font_c, fill_c = TONE_BAD
    elif days is not None and days <= 30:
        font_c, fill_c = TONE_WARN
    elif label == "Sin rotación":
        font_c, fill_c = TONE_NEUTRAL
    else:
        font_c, fill_c = None, None  # normal: hereda color de fila

    if days is not None:
        return {"value": days, "is_number": True, "font": font_c, "fill": fill_c}
    return {"value": label or "—", "is_number": False, "font": font_c, "fill": fill_c}


# Formato para la columna de días (número + sufijo)
NF_DAYS = '0" días"'


# ──────────────────────────────────────────────────────────────────────────
# Helpers de número/fecha/sanitización
# ──────────────────────────────────────────────────────────────────────────

def _num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _safe_sheet_name(name: str, used: set[str], max_len: int = 31) -> str:
    """Excel: máx 31 chars, sin / \\ ? * [ ] :"""
    cleaned = "".join(c for c in name if c not in r'/\?*[]:')[:max_len] or "Hoja"
    base = cleaned
    i = 2
    while cleaned in used:
        suf = f" ({i})"
        cleaned = (base[: max_len - len(suf)] + suf)
        i += 1
    used.add(cleaned)
    return cleaned


def _safe_date(v: Any):
    """Devuelve datetime/date o '—' para celdas vacías."""
    if v in (None, "", "—"):
        return None
    if isinstance(v, (datetime, date)):
        return v
    s = str(v).strip()
    if not s:
        return None
    # ISO o YYYY-MM-DD
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[: len(fmt) + 6], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


# ──────────────────────────────────────────────────────────────────────────
# Aggregation: rows[] (dict por columna) → árbol Depto → Cat → SubCat → SKU
# Si la matriz trae varias sucursales por SKU (Sucursal = "Todas"), agrupa
# por SKU sumando ventas/unds/stock, min de 1ª recep, max de Últ Venta, y
# elige la clasificación de la fila con más ventas.
# ──────────────────────────────────────────────────────────────────────────

def _aggregate_rows(rows_dict: list[dict[str, Any]]) -> dict:
    """Devuelve {dept: {cat: {subcat: {sku: {**fields}}}}}."""
    tree: dict[str, dict[str, dict[str, dict[str, dict[str, Any]]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(dict))
    )

    for r in rows_dict:
        dept   = str(r.get("Departamento") or "—")
        cat    = str(r.get("Categoría") or "—")
        subcat = str(r.get("Subcategoría") or "—")
        sku    = str(r.get("Código SKU") or "—")

        agg = tree[dept][cat][subcat].get(sku)
        ventas = _num(r.get("Vendido SKU S/"))
        # ★ "Vend Lote" = unidades vendidas del LOTE actual completo (desde
        #   `1ª Venta Lote` hasta hoy), NO los últimos 90d. Razón: para SKUs
        #   cuyo lote es viejo (>90d), la cifra de 90d es engañosa — un
        #   producto que llegó hace 130d, vendió 274 y se agotó, mostraría
        #   "9" si sólo miramos 90d. La cifra del lote refleja el éxito real
        #   del último ciclo (que es lo que pregunta el dueño del negocio).
        unds = _num(r.get("Vend Lote Total"))
        stock = _num(r.get("Stock Disp"))
        clasif = str(r.get("Clasificación") or "")
        tend = str(r.get("Tendencia") or "")
        prod = str(r.get("Producto") or "")
        # ★ "Ingreso" = ÚLTIMA recepción (cuándo llegó el último lote), no la
        #   primera. Es lo operativamente útil: para evaluar reposición importa
        #   "¿hace cuánto entró stock nuevo?", no "¿cuándo nació el SKU?".
        #   Cuando el SKU está en varias sucursales, tomamos la más RECIENTE
        #   (MAX), igual criterio que con la última venta.
        last_recep = _safe_date(r.get("Últ. Recepción"))
        last_sale = _safe_date(r.get("Fecha Últ. Venta"))
        # ★ "Días Stock" = cuántos días de inventario le quedan al SKU al ritmo
        #   de venta actual. La matriz ya lo calcula como columna "Cobertura"
        #   (texto: "110 días" / "Agotado") usando la velocidad del lote actual.
        #   Guardamos el TEXTO del SQL (para reusarlo tal cual cuando el SKU
        #   está en una sola sucursal) y la velocidad (uds/día) para poder
        #   RECALCULAR los días cuando se consolidan varias sucursales
        #   (stock_total ÷ velocidad_total).
        vel = _num(r.get("Velocidad (uds/día)"))
        cobertura_txt = str(r.get("Cobertura") or "")
        # ★ P18/P19 (2026-06-08): nuevas métricas informativas. Vienen del SQL.
        llego_hace = r.get("Llegó hace (días)")  # int o None
        sell_through = r.get("Sell-through Lote %")  # float (0-100) o None
        vida_lote = r.get("Vida lote (días)")  # int o None
        # ★ Días Exhibido (2026-06-11): días reales con stock>0 en tienda.
        dias_exhibido = r.get("Días Exhibido")  # int o None
        # ★ P22: stock almacén es por VARIANTE (misma cifra en ambas sucursales)
        #   → al consolidar se toma el MAX, no la suma (evita duplicar).
        stock_alm = _num(r.get("Stock Almacén"))

        if agg is None:
            agg = {
                "producto": prod,
                "ventas_s": ventas,
                "unds": unds,
                "stock": stock,
                "velocidad": vel,
                "cobertura_txt": cobertura_txt,
                "clasif": clasif,
                "tendencia": tend,
                "ingreso": last_recep,
                "ult_venta": last_sale,
                # Métricas informativas — tomamos las del dominante (ver abajo).
                "llego_hace": llego_hace,
                "sell_through": sell_through,
                "vida_lote": vida_lote,
                "dias_exhibido": dias_exhibido,
                "stock_almacen": stock_alm,
                "_dominant_ventas": ventas,  # para elegir clasif dominante
                "_n_filas": 1,               # cuántas sucursales se consolidaron
            }
            tree[dept][cat][subcat][sku] = agg
        else:
            agg["ventas_s"] += ventas
            agg["unds"] += unds
            agg["stock"] += stock
            agg["velocidad"] += vel
            agg["stock_almacen"] = max(agg.get("stock_almacen", 0.0), stock_alm)
            agg["_n_filas"] += 1
            # Max ingreso (última recep más reciente entre sucursales), Max últ venta
            if last_recep and (agg["ingreso"] is None or last_recep > agg["ingreso"]):
                agg["ingreso"] = last_recep
            if last_sale and (agg["ult_venta"] is None or last_sale > agg["ult_venta"]):
                agg["ult_venta"] = last_sale
            # Clasificación + métricas de la sucursal con MÁS ventas (dominante).
            # Para "Llegó hace", "Sell-through" y "Vida lote" tiene sentido tomar
            # las del lote dominante porque representan el lote más vendido.
            if ventas > agg["_dominant_ventas"]:
                agg["_dominant_ventas"] = ventas
                agg["clasif"] = clasif
                agg["tendencia"] = tend
                agg["cobertura_txt"] = cobertura_txt
                agg["llego_hace"] = llego_hace
                agg["sell_through"] = sell_through
                agg["vida_lote"] = vida_lote
                agg["dias_exhibido"] = dias_exhibido

    return tree


def _dominant_label(skus: dict[str, dict]) -> str:
    """Clasificación dominante de una subcat: la más frecuente entre sus SKUs.
    Si hay empate, gana la primera por orden alfabético (determinista)."""
    counts = Counter(s["clasif"].split(":")[0].strip() for s in skus.values() if s["clasif"])
    if not counts:
        return ""
    # Counter.most_common no es determinista en empates → ordeno explícito
    best_count = max(counts.values())
    candidates = sorted(k for k, v in counts.items() if v == best_count)
    return candidates[0]


# ──────────────────────────────────────────────────────────────────────────
# Apply cell helper — encapsula font + fill + align + number_format + border
# ──────────────────────────────────────────────────────────────────────────

def _apply(cell, *, font: Font | None = None, fill: PatternFill | None = None,
           align: Alignment | None = None, num_format: str | None = None,
           border: Border | None = None) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if num_format is not None:
        cell.number_format = num_format
    if border is not None:
        cell.border = border


def _row_fill(ws: Worksheet, row: int, fill: PatternFill, col_count: int) -> None:
    """Aplica el mismo fill a TODAS las celdas de la fila (col 1..col_count)."""
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).fill = fill


# ──────────────────────────────────────────────────────────────────────────
# Resumen (1er hoja): ranking de departamentos
# ──────────────────────────────────────────────────────────────────────────

# ★ Columnas de las hojas por departamento, agrupadas por tema para que el
#   análisis se lea de izquierda a derecha: QUIÉN es el producto → cuánto
#   VENDE → qué STOCK tiene → su CICLO de vida en tienda → el DIAGNÓSTICO.
#   (Reordenadas 2026-06-11; antes las informativas quedaban apiladas al
#   final en orden cronológico de cuando se fueron agregando.)
DEPT_HEADERS = [
    # ── Identificación ──
    "Tipo", "Categoría", "Subcategoría", "SKU", "Producto",
    # ── Ventas (qué tan bien vende) ──
    "Ventas (S/)", "% Total", "Vend Lote", "Sell-through %",
    # ── Stock (qué hay hoy y cuánto dura) ──
    #   Stock Almacén (P22): si >0 y la tienda está baja → TRASLADAR, no comprar.
    "Stock", "Stock Almacén", "Días Stock",
    # ── Ciclo del lote (fechas y exhibición) ──
    "Ingreso", "Llegó hace (días)", "Días Exhibido", "Últ. Venta", "Vida lote (días)",
    # ── Diagnóstico (la conclusión, al final) ──
    "Tendencia", "Clasificación",
]

# Índices 1-based derivados de DEPT_HEADERS: el orden se define UNA vez en la
# lista de arriba; reordenarla no requiere tocar los writers de filas.
_COL = {h: i + 1 for i, h in enumerate(DEPT_HEADERS)}
COL_TIPO        = _COL["Tipo"]
COL_CAT         = _COL["Categoría"]
COL_SUBCAT      = _COL["Subcategoría"]
COL_SKU         = _COL["SKU"]
COL_PROD        = _COL["Producto"]
COL_VENTAS      = _COL["Ventas (S/)"]
COL_PCT         = _COL["% Total"]
COL_VEND_LOTE   = _COL["Vend Lote"]
COL_SELLTHROUGH = _COL["Sell-through %"]
COL_STOCK       = _COL["Stock"]
COL_STOCK_ALM   = _COL["Stock Almacén"]
COL_DIAS_STOCK  = _COL["Días Stock"]
COL_INGRESO     = _COL["Ingreso"]
COL_LLEGO       = _COL["Llegó hace (días)"]
COL_DIAS_EXHIB  = _COL["Días Exhibido"]
COL_ULT_VENTA   = _COL["Últ. Venta"]
COL_VIDA_LOTE   = _COL["Vida lote (días)"]
COL_TENDENCIA   = _COL["Tendencia"]
COL_CLASIF      = _COL["Clasificación"]


def _build_resumen(wb: Workbook, tree: dict, meta_line: str, total_general: float,
                   counts: dict[str, int]) -> dict[str, float]:
    """Crea la hoja Resumen y devuelve el ranking {dept: ventas} para los tabs."""
    ws = wb.active
    ws.title = "📊 Resumen"
    ws.sheet_properties.tabColor = C_TITLE_BG

    # Anchos de columna
    widths = [6, 38, 18, 13, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title (rosa→naranja)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    c = ws.cell(row=1, column=1, value="🏷️ REPORTE DE VENTAS HUDEC")
    _apply(c, font=_font(C_TITLE_FG, size=20, bold=True),
           fill=_fill(C_TITLE_BG), align=_align("left", "center"))
    ws.row_dimensions[1].height = 34

    # Meta
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    c = ws.cell(row=2, column=1, value=meta_line)
    _apply(c, font=_font(C_META_FG, size=10),
           fill=_fill(C_META_BG), align=_align("left", "center", wrap=True))
    ws.row_dimensions[2].height = 22

    # Total del período
    c = ws.cell(row=3, column=1, value="Total del período")
    _apply(c, font=_font(C_TOTAL_LBL, size=11, bold=True),
           fill=_fill(C_TOTAL_BG), align=_align("left", "center"))
    c = ws.cell(row=3, column=2, value=total_general)
    _apply(c, font=_font(C_TOTAL_FG, size=18, bold=True),
           fill=_fill(C_TOTAL_BG), align=_align("left", "center"),
           num_format=NF_MONEY)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)
    note = (f"{counts.get('depts', 0)} departamentos · "
            f"{counts.get('cats', 0)} categorías · "
            f"{counts.get('subcats', 0)} subcategorías")
    c = ws.cell(row=3, column=3, value=note)
    _apply(c, font=_font(C_NOTE_FG, size=9, italic=True),
           fill=_fill(C_TOTAL_BG), align=_align("left", "center", wrap=True))
    ws.row_dimensions[3].height = 36

    # Headers del ranking
    headers = ["#", "Departamento", "Ventas (S/)", "% del Total", "Tickets", "Ticket Prom."]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=i, value=h)
        _apply(c, font=_font(C_HEADER_FG, size=10, bold=True),
               fill=_fill(C_HEADER_BG), align=_align("center", "center"),
               border=_border_bottom("000000"))
    ws.row_dimensions[5].height = 22

    # Ranking de departamentos por ventas descendente
    dept_totals = [
        (d, sum(sku["ventas_s"]
                for cat in tree[d].values()
                for sub in cat.values()
                for sku in sub.values()),
         sum(sku["unds"]
             for cat in tree[d].values()
             for sub in cat.values()
             for sku in sub.values()))
        for d in tree
    ]
    dept_totals.sort(key=lambda x: x[1], reverse=True)

    rank_for_dept: dict[str, float] = {}
    row = 6
    for rank, (dept, ventas, tickets) in enumerate(dept_totals, start=1):
        rank_for_dept[dept] = ventas
        # Rank
        c = ws.cell(row=row, column=1, value=rank)
        _apply(c, font=_font(C_RANK_FG, size=11, bold=True),
               fill=_fill(C_RANK_BG), align=_align("center"))
        # Dept
        c = ws.cell(row=row, column=2, value=dept)
        _apply(c, font=_font(C_DEPT_FG, size=13, bold=True),
               fill=_fill(C_DEPT_BG), align=_align("left"))
        # Ventas S/
        c = ws.cell(row=row, column=3, value=ventas)
        _apply(c, font=_font(C_DEPT_FG, size=13, bold=True),
               fill=_fill(C_DEPT_BG), align=_align("right"),
               num_format=NF_MONEY)
        # %
        pct = (ventas / total_general) if total_general else 0
        c = ws.cell(row=row, column=4, value=pct)
        _apply(c, font=_font(C_DEPT_FG, size=13, bold=True),
               fill=_fill(C_DEPT_BG), align=_align("right"),
               num_format=NF_PCT)
        # Tickets/unds
        c = ws.cell(row=row, column=5, value=tickets)
        _apply(c, font=_font(C_DEPT_FG, size=13, bold=True),
               fill=_fill(C_DEPT_BG), align=_align("right"),
               num_format=NF_INT)
        # Ticket prom
        avg = (ventas / tickets) if tickets else 0
        c = ws.cell(row=row, column=6, value=avg)
        _apply(c, font=_font(C_DEPT_FG, size=13, bold=True),
               fill=_fill(C_DEPT_BG), align=_align("right"),
               num_format=NF_MONEY)
        ws.row_dimensions[row].height = 22
        row += 1

    # Freeze los encabezados
    ws.freeze_panes = "A6"
    return rank_for_dept


# ──────────────────────────────────────────────────────────────────────────
# Hoja por Departamento — cuerpo plano jerárquico
# ──────────────────────────────────────────────────────────────────────────

# Ancho por NOMBRE de columna (robusto a reordenamientos de DEPT_HEADERS).
_WIDTH_BY_HEADER = {
    "Tipo": 14, "Categoría": 22, "Subcategoría": 22, "SKU": 14, "Producto": 34,
    "Ventas (S/)": 14, "% Total": 10, "Vend Lote": 12, "Sell-through %": 13,
    "Stock": 9, "Stock Almacén": 12, "Días Stock": 11,
    "Ingreso": 12, "Llegó hace (días)": 12, "Días Exhibido": 12,
    "Últ. Venta": 12, "Vida lote (días)": 13,
    "Tendencia": 16, "Clasificación": 38,
}
DEPT_COL_WIDTHS = [_WIDTH_BY_HEADER[h] for h in DEPT_HEADERS]


def _build_dept_sheet(wb: Workbook, dept_name: str, dept_data: dict,
                      rank: int, total_dept: float, total_general: float,
                      meta_line: str, used_names: set[str], tab_color: str) -> None:
    sheet_name = _safe_sheet_name(f"{rank}. {dept_name}", used_names)
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    NCOLS = len(DEPT_HEADERS)

    # Anchos
    for i, w in enumerate(DEPT_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    c = ws.cell(row=1, column=1, value=f"🏷️ {dept_name}")
    _apply(c, font=_font(C_TITLE_FG, size=20, bold=True),
           fill=_fill(C_TITLE_BG), align=_align("left"))
    _row_fill(ws, 1, _fill(C_TITLE_BG), NCOLS)
    ws.row_dimensions[1].height = 30

    # Meta
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    c = ws.cell(row=2, column=1, value=f"Ranking #{rank} · {meta_line}")
    _apply(c, font=_font(C_META_FG, size=10),
           fill=_fill(C_META_BG), align=_align("left", "center", wrap=True))
    _row_fill(ws, 2, _fill(C_META_BG), NCOLS)
    ws.row_dimensions[2].height = 20

    # Total dept
    total_unds = sum(sku["unds"] for cat in dept_data.values()
                     for sub in cat.values() for sku in sub.values())
    sku_count = sum(len(sub) for cat in dept_data.values() for sub in cat.values())
    pct_total = (total_dept / total_general) if total_general else 0

    c = ws.cell(row=3, column=1, value="Total departamento")
    _apply(c, font=_font(C_TOTAL_LBL, size=11, bold=True),
           fill=_fill(C_TOTAL_BG), align=_align("left"))
    c = ws.cell(row=3, column=2, value=total_dept)
    _apply(c, font=_font(C_TOTAL_FG, size=18, bold=True),
           fill=_fill(C_TOTAL_BG), align=_align("left"),
           num_format=NF_MONEY)
    note = (f"{pct_total * 100:.1f}% del total · "
            f"{total_unds:,.0f} unds · {sku_count} SKUs")
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=NCOLS)
    c = ws.cell(row=3, column=3, value=note)
    _apply(c, font=_font(C_NOTE_FG, size=9, italic=True),
           fill=_fill(C_TOTAL_BG), align=_align("left", "center", wrap=True))
    _row_fill(ws, 3, _fill(C_TOTAL_BG), NCOLS)
    ws.row_dimensions[3].height = 32

    # Headers (fila 5; la 4 queda en blanco)
    header_row = 5
    for i, h in enumerate(DEPT_HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        _apply(c, font=_font(C_HEADER_FG, size=10, bold=True),
               fill=_fill(C_HEADER_BG), align=_align("center", "center", wrap=True),
               border=_border_bottom("000000"))
    ws.row_dimensions[header_row].height = 24

    # Cuerpo: Cat → SubCat → Producto
    row = header_row + 1
    alt = False

    # Categorías ordenadas por ventas desc
    cats_sorted = sorted(
        dept_data.items(),
        key=lambda kv: sum(s["ventas_s"]
                           for sub in kv[1].values() for s in sub.values()),
        reverse=True,
    )

    for cat_name, cat_data in cats_sorted:
        cat_ventas = sum(s["ventas_s"] for sub in cat_data.values() for s in sub.values())
        cat_unds = sum(s["unds"] for sub in cat_data.values() for s in sub.values())
        cat_pct = (cat_ventas / total_general) if total_general else 0
        # ★ Veredicto AGREGADO de la categoría (mix de ventas de todos sus SKUs)
        cat_skus = aggregate_cat_skus(cat_data)
        cat_ver = compute_veredicto(cat_skus)
        # Contar veredictos de subcats: cuántas 🟢/🟡/🔴
        sub_verdicts = [compute_veredicto(sd) for sd in cat_data.values()]
        n_verde = sum(1 for v in sub_verdicts if "SEGUIR" in v["veredicto"])
        n_amari = sum(1 for v in sub_verdicts if "EVALUAR" in v["veredicto"])
        n_rojo  = sum(1 for v in sub_verdicts if "REDUCIR" in v["veredicto"])

        # Fila Cat (rosa→naranja claro, bold)
        cat_fill = _fill(C_CAT_BG)
        _row_fill(ws, row, cat_fill, NCOLS)
        c = ws.cell(row=row, column=1, value="📂 Categoría")
        _apply(c, font=_font(C_CAT_FG, size=11, bold=True), fill=cat_fill, align=_align("left"))
        c = ws.cell(row=row, column=2, value=cat_name)
        _apply(c, font=_font(C_CAT_FG, size=11, bold=True), fill=cat_fill, align=_align("left"))
        c = ws.cell(row=row, column=COL_VENTAS, value=cat_ventas)
        _apply(c, font=_font(C_CAT_FG, size=11, bold=True), fill=cat_fill,
               align=_align("right"), num_format=NF_MONEY)
        c = ws.cell(row=row, column=COL_PCT, value=cat_pct)
        _apply(c, font=_font(C_CAT_FG, size=11, bold=True), fill=cat_fill,
               align=_align("right"), num_format=NF_PCT)
        c = ws.cell(row=row, column=COL_VEND_LOTE, value=cat_unds)
        _apply(c, font=_font(C_CAT_FG, size=11, bold=True), fill=cat_fill,
               align=_align("right"), num_format=NF_INT)
        # ★ Veredicto agregado de la Cat (columna Clasificación)
        c = ws.cell(row=row, column=COL_CLASIF, value=cat_ver["veredicto"])
        _apply(c, font=_font(cat_ver["tone"][0], size=11, bold=True),
               fill=_fill(cat_ver["tone"][1]), align=_align("left", "center", wrap=True))
        # ★ Conteo de subcats por veredicto (columna Tendencia)
        sub_summary = f"{n_verde}🟢 · {n_amari}🟡 · {n_rojo}🔴"
        c = ws.cell(row=row, column=COL_TENDENCIA, value=sub_summary)
        _apply(c, font=_font(C_CAT_FG, size=9, italic=True), fill=cat_fill, align=_align("left"))
        ws.row_dimensions[row].height = 22
        row += 1

        # Subcategorías ordenadas por ventas desc
        subs_sorted = sorted(
            cat_data.items(),
            key=lambda kv: sum(s["ventas_s"] for s in kv[1].values()),
            reverse=True,
        )

        for sub_name, sub_data in subs_sorted:
            sub_ventas = sum(s["ventas_s"] for s in sub_data.values())
            sub_unds = sum(s["unds"] for s in sub_data.values())
            sub_pct = (sub_ventas / total_general) if total_general else 0
            # ★ Veredicto de la SubCat: ¿Seguir comprando o descartar?
            sub_ver = compute_veredicto(sub_data)
            sub_descartar = "REDUCIR" in sub_ver["veredicto"]

            sub_fill = _fill(C_SUBCAT_BG)
            _row_fill(ws, row, sub_fill, NCOLS)
            c = ws.cell(row=row, column=1, value="📁 Subcategoría")
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill, align=_align("left"))
            c = ws.cell(row=row, column=2, value=cat_name)
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill, align=_align("left"))
            c = ws.cell(row=row, column=3, value=sub_name)
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill, align=_align("left"))
            c = ws.cell(row=row, column=COL_VENTAS, value=sub_ventas)
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill,
                   align=_align("right"), num_format=NF_MONEY)
            c = ws.cell(row=row, column=COL_PCT, value=sub_pct)
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill,
                   align=_align("right"), num_format=NF_PCT)
            c = ws.cell(row=row, column=COL_VEND_LOTE, value=sub_unds)
            _apply(c, font=_font(C_SUBCAT_FG, size=10, bold=True), fill=sub_fill,
                   align=_align("right"), num_format=NF_INT)
            # ★ Veredicto de la SubCat (columna Clasificación)
            c = ws.cell(row=row, column=COL_CLASIF, value=sub_ver["veredicto"])
            _apply(c, font=_font(sub_ver["tone"][0], size=10, bold=True),
                   fill=_fill(sub_ver["tone"][1]), align=_align("left", "center", wrap=True))
            # ★ Razón + conteo (columna Tendencia)
            parts = [sub_ver["razon"]]
            if sub_ver["n_estrellas"]:
                parts.append(f"{sub_ver['n_estrellas']}⭐")
            if sub_ver["n_problemas_stock"]:
                parts.append(f"{sub_ver['n_problemas_stock']}💸 con stock")
            note_txt = " · ".join(parts)
            c = ws.cell(row=row, column=COL_TENDENCIA, value=note_txt)
            _apply(c, font=_font(C_SUBCAT_FG, size=9, italic=True),
                   fill=sub_fill, align=_align("left", "center", wrap=True))
            ws.row_dimensions[row].height = 22
            row += 1

            # Productos ordenados por ventas desc
            skus_sorted = sorted(
                sub_data.items(),
                key=lambda kv: kv[1]["ventas_s"], reverse=True,
            )

            for sku, info in skus_sorted:
                # ★ Rescatable redefinido: SKU SANO (bucket comprable) que vive
                #   en una SubCat marcada 🔴 REDUCIR/DESCARTAR. No se debe perder.
                prod_bucket = categorize_bucket(info["clasif"])
                rescatable = sub_descartar and prod_bucket == "comprable"

                if rescatable:
                    bg = _fill(C_RESCATE_BG)
                    fg = C_RESCATE_FG
                    bold = True
                    border = _border_left_thick(C_RESCATE_BORDER)
                else:
                    bg = _fill(C_PROD_BG_ALT if alt else C_PROD_BG)
                    fg = C_PROD_FG
                    bold = False
                    border = None
                alt = not alt

                _row_fill(ws, row, bg, NCOLS)
                c = ws.cell(row=row, column=1, value=("⭐ Rescatable" if rescatable else "▸ Producto"))
                _apply(c, font=_font(fg, size=10, bold=bold), fill=bg, align=_align("left"), border=border)
                c = ws.cell(row=row, column=2, value=cat_name)
                _apply(c, font=_font(fg, size=10, bold=bold), fill=bg, align=_align("left"))
                c = ws.cell(row=row, column=3, value=sub_name)
                _apply(c, font=_font(fg, size=10, bold=bold), fill=bg, align=_align("left"))
                c = ws.cell(row=row, column=4, value=sku)
                _apply(c, font=_font(C_PROD_MONO, size=9, name="Consolas"),
                       fill=bg, align=_align("left"))
                c = ws.cell(row=row, column=5, value=info["producto"])
                _apply(c, font=_font(fg, size=10, bold=bold), fill=bg, align=_align("left"))
                c = ws.cell(row=row, column=COL_VENTAS, value=info["ventas_s"])
                _apply(c, font=_font(fg, size=10, bold=bold), fill=bg,
                       align=_align("right"), num_format=NF_MONEY)
                pct_sku = (info["ventas_s"] / total_general) if total_general else 0
                c = ws.cell(row=row, column=COL_PCT, value=pct_sku)
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format=NF_PCT)
                c = ws.cell(row=row, column=COL_VEND_LOTE, value=info["unds"])
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format=NF_INT)
                # Sell-through Lote %
                st = info.get("sell_through")
                c = ws.cell(row=row, column=COL_SELLTHROUGH, value=float(st)/100 if st is not None else None)
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format="0.0%")
                c = ws.cell(row=row, column=COL_STOCK, value=info["stock"])
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format=NF_INT)
                # ★ P22 — Stock Almacén. Azul si >0 (hay backup:
                #   la acción es TRASLADAR, no comprar al proveedor).
                alm = info.get("stock_almacen", 0.0)
                c = ws.cell(row=row, column=COL_STOCK_ALM, value=int(alm) if alm > 0 else None)
                if alm > 0:
                    _apply(c, font=_font("1F4E79", size=10, bold=True),
                           fill=_fill("DDEBF7"), align=_align("right"), num_format=NF_INT)
                else:
                    _apply(c, font=_font(fg, size=10), fill=bg,
                           align=_align("right"), num_format=NF_INT)
                # Días Stock (cobertura: días de inventario que le quedan al SKU)
                cov = _coverage_cell(
                    info["stock"], info.get("velocidad", 0.0),
                    info.get("cobertura_txt", ""),
                    merged=info.get("_n_filas", 1) > 1,
                )
                c = ws.cell(row=row, column=COL_DIAS_STOCK, value=cov["value"])
                _apply(
                    c,
                    font=_font(cov["font"] or fg, size=10, bold=bold or cov["font"] is not None),
                    fill=_fill(cov["fill"]) if cov["fill"] else bg,
                    align=_align("right"),
                    num_format=NF_DAYS if cov["is_number"] else None,
                )
                # Ingreso (última recepción)
                if info["ingreso"]:
                    c = ws.cell(row=row, column=COL_INGRESO, value=info["ingreso"])
                    _apply(c, font=_font(fg, size=10), fill=bg,
                           align=_align("center"), num_format=NF_DATE)
                else:
                    c = ws.cell(row=row, column=COL_INGRESO, value="—")
                    _apply(c, font=_font(fg, size=10), fill=bg, align=_align("center"))
                # Llegó hace (días)
                llego = info.get("llego_hace")
                c = ws.cell(row=row, column=COL_LLEGO, value=int(llego) if llego is not None else None)
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format=NF_DAYS)
                # ★ Días Exhibido (2026-06-11): días reales con stock>0 en tienda
                #   dentro del ciclo del lote. Complementa "Llegó hace": un lote
                #   puede llevar 60d en tienda pero solo 20 exhibido (resto agotado).
                exhib = info.get("dias_exhibido")
                c = ws.cell(row=row, column=COL_DIAS_EXHIB, value=int(exhib) if exhib is not None else None)
                _apply(c, font=_font(fg, size=10), fill=bg,
                       align=_align("right"), num_format=NF_DAYS)
                # Últ Venta
                if info["ult_venta"]:
                    c = ws.cell(row=row, column=COL_ULT_VENTA, value=info["ult_venta"])
                    _apply(c, font=_font(fg, size=10), fill=bg,
                           align=_align("center"), num_format=NF_DATE)
                else:
                    c = ws.cell(row=row, column=COL_ULT_VENTA, value="—")
                    _apply(c, font=_font(fg, size=10), fill=bg, align=_align("center"))
                # Vida lote (días) — con tinte visual según rangos
                vida = info.get("vida_lote")
                c = ws.cell(row=row, column=COL_VIDA_LOTE, value=int(vida) if vida is not None else None)
                vida_font, vida_fill = fg, bg
                if vida is not None:
                    if vida > 150:
                        vida_font = "9C0006"  # rojo (lote sobredimensionado)
                        vida_fill = _fill("FFC7CE")
                    elif vida > 90:
                        vida_font = "9C5700"  # ámbar
                        vida_fill = _fill("FFEB9C")
                _apply(c, font=_font(vida_font, size=10,
                       bold=vida is not None and vida > 90),
                       fill=vida_fill if vida is not None and vida > 90 else bg,
                       align=_align("right"), num_format=NF_DAYS)
                # Tendencia
                c = ws.cell(row=row, column=COL_TENDENCIA, value=info["tendencia"] or "—")
                _apply(c, font=_font(fg, size=10), fill=bg, align=_align("left"))
                # Clasificación con tone (la conclusión, última columna)
                tone = classify_tone(info["clasif"])
                c = ws.cell(row=row, column=COL_CLASIF, value=info["clasif"] or "—")
                _apply(c, font=_font(tone[0], size=10, bold=True),
                       fill=_fill(tone[1]), align=_align("left", "center", wrap=True))
                ws.row_dimensions[row].height = 18
                row += 1

    # Freeze: header + columnas de identificación (Tipo..Producto), para que
    # al hacer scroll horizontal siempre se vea QUÉ producto es cada fila.
    ws.freeze_panes = f"{get_column_letter(COL_PROD + 1)}{header_row + 1}"


# ──────────────────────────────────────────────────────────────────────────
# Entry point — drop-in para excel_builder.build_workbook
# ──────────────────────────────────────────────────────────────────────────

def build_executive_workbook(
    *,
    cols: list[str],
    rows: Iterable[tuple],
    modulo_id: str,
    titulo: str = "Reporte Ventas HUDEC",
    sql_file: str = "",
    descripcion: str = "",
    classification_col: str = "Clasificación",
    elapsed_seconds: float | None = None,
    brand_name: str = "hudec",
    # Filtros aplicados (van al subtítulo Meta)
    sucursal: str | None = None,
    accion_label: str | None = None,
    periodo_dias: int = 60,
) -> Workbook:
    """Construye el workbook ejecutivo (formato COYA-naranja).

    Mantiene la misma firma genérica que `excel_builder.build_workbook` para que
    el router pueda elegir cuál llamar. Los kwargs `sucursal`/`accion_label`/etc.
    son opcionales y van al meta-subtítulo de cada hoja.
    """
    # rows (Iterable[tuple]) → list[dict] usando cols
    rows_dict: list[dict[str, Any]] = []
    for r in rows:
        rows_dict.append({c: r[i] for i, c in enumerate(cols) if i < len(r)})

    tree = _aggregate_rows(rows_dict)

    # Total general
    total_general = sum(
        sku["ventas_s"]
        for d in tree.values()
        for c in d.values()
        for sub in c.values()
        for sku in sub.values()
    )

    counts = {
        "depts": len(tree),
        "cats": sum(len(d) for d in tree.values()),
        "subcats": sum(len(c) for d in tree.values() for c in d.values()),
    }

    now = datetime.now()
    meta_parts = [f"📅 Últimos {periodo_dias} días"]
    meta_parts.append(f"🏪 {sucursal}" if sucursal else "🏪 Todas las sucursales")
    meta_parts.append(f"🎯 {accion_label}" if accion_label else "🎯 Todos")
    meta_parts.append(f"🕐 {now.strftime('%d/%m/%Y, %I:%M %p').lower()}")
    meta_line = "  ·  ".join(meta_parts)

    wb = Workbook()
    rank_for_dept = _build_resumen(wb, tree, meta_line, total_general, counts)

    # Hojas por dept (orden = ranking)
    used_names = {wb.active.title}
    sorted_depts = sorted(tree.keys(), key=lambda d: rank_for_dept.get(d, 0), reverse=True)
    for rank, dept_name in enumerate(sorted_depts, start=1):
        total_dept = rank_for_dept.get(dept_name, 0)
        tab_color = TAB_COLORS[(rank - 1) % len(TAB_COLORS)]
        _build_dept_sheet(
            wb, dept_name, tree[dept_name], rank, total_dept,
            total_general, meta_line, used_names, tab_color,
        )

    return wb
