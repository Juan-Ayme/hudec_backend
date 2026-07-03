"""Builder Excel del **Informe Diario** de gerencia (mes en curso).

3 pestañas exclusivas, todas centradas en el seguimiento día a día del mes actual:
    Ticket promedio · Transacciones · Venta vs meta (diario)

Diferencias con el tablero semanal:
    - Solo cubre el mes en curso (no acepta otros meses).
    - 'Venta vs meta' es DIARIA (una fila por día del mes), no mensual.
    - Cada hoja incluye su gráfico de línea/barras al costado.

Consume el dict que devuelve `app.routers.analytics.daily_report` y reutiliza
los estilos compartidos de `analytics.excel_builder`.
"""

from __future__ import annotations

from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analytics.excel_builder import (
    BORDER_THIN,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    ancho_columna,
)

# ── Formatos de celda ──────────────────────────────────────────────────────
MONEY_FMT = '"S/ "#,##0.00'
PCT_FMT = '0.0"%"'
INT_FMT = "#,##0"

TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
LABEL_FONT = Font(bold=True)
NOTE_FONT = Font(italic=True, color="808080")

TAB_TICKET = "2E75B6"
TAB_TRX = "385723"
TAB_META = "BF8F00"

FILL_OK = PatternFill("solid", fgColor="C8E6C9")
FILL_BAD = PatternFill("solid", fgColor="FFB3B3")
FILL_FUTURO = PatternFill("solid", fgColor="F0F0F0")


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _title(ws: Worksheet, text: str, ncols: int) -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: list[Sequence[Any]],
    *,
    money_cols: set[int] = frozenset(),
    pct_cols: set[int] = frozenset(),
    int_cols: set[int] = frozenset(),
    start_row: int = 2,
) -> int:
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
    ws.row_dimensions[start_row].height = 26

    r = start_row + 1
    for row_vals in rows:
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = BORDER_THIN
            col0 = c_idx - 1
            if col0 in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col0 in pct_cols:
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col0 in int_cols:
                cell.number_format = INT_FMT
                cell.alignment = Alignment(horizontal="right")
        r += 1

    last = r - 1
    for c_idx, h in enumerate(headers, 1):
        values = [row_vals[c_idx - 1] for row_vals in rows[:200]] if rows else []
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho_columna(values, h)
    ws.freeze_panes = f"A{start_row + 1}"
    if last >= start_row + 1:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last}"
    return last


# ── Hojas ──────────────────────────────────────────────────────────────────
def _sheet_ticket(wb: Workbook, daily: dict) -> None:
    """Ticket promedio diario del mes — tabla + LineChart."""
    ws = wb.create_sheet("Ticket promedio")
    ws.sheet_properties.tabColor = TAB_TICKET
    _title(ws, f"Ticket promedio diario — {daily.get('month', '')}", 2)
    serie = daily.get("serie", [])
    rows = [(p.get("fecha"), _num(p.get("ticket_promedio"))) for p in serie]
    last = _write_table(ws, ["Día", "Ticket promedio (S/)"], rows, money_cols={1})
    if last >= 3:
        chart = LineChart()
        chart.title = f"Ticket promedio (S/) por día — {daily.get('month', '')}"
        chart.y_axis.title = "Ticket promedio (S/)"
        chart.x_axis.title = "Fecha"
        chart.height = 10
        chart.width = 22
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=last, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D2")


def _sheet_transacciones(wb: Workbook, daily: dict) -> None:
    """N° transacciones por día del mes — tabla + BarChart."""
    ws = wb.create_sheet("Transacciones")
    ws.sheet_properties.tabColor = TAB_TRX
    _title(ws, f"N° de transacciones por día — {daily.get('month', '')}", 2)
    serie = daily.get("serie", [])
    rows = [(p.get("fecha"), _num(p.get("tickets"))) for p in serie]
    last = _write_table(ws, ["Día", "N° de transacciones"], rows, int_cols={1})
    if last >= 3:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"N° de transacciones por día — {daily.get('month', '')}"
        chart.y_axis.title = "Transacciones"
        chart.x_axis.title = "Fecha"
        chart.height = 10
        chart.width = 22
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=last, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D2")


def _sheet_meta_diaria(wb: Workbook, daily: dict) -> None:
    """Venta vs meta — 4 bloques verticales con vista por sucursal.

    Bloque 1: 🎯 Desglose de metas diarias (tabla por sucursal)
    Bloque 2: 📊 Reporte diario (indicadores por sucursal al día de hoy)
    Bloque 3: ✅ Semáforo (regla operativa)
    Bloque 4: 📋 Tabla diaria de seguimiento (una fila por día del mes)
    """
    ws = wb.create_sheet("Venta vs meta")
    ws.sheet_properties.tabColor = TAB_META
    dias_mes = int(daily.get("dias_del_mes") or 0)
    dias_t = int(daily.get("dias_transcurridos") or 0)
    sucursales = daily.get("sucursales") or []
    _title(
        ws,
        f"Venta vs meta — {daily.get('month', '')} · corte día {dias_t}/{dias_mes}",
        7,
    )

    cursor = 3  # primera fila libre debajo del título
    cursor = _block_desglose_metas(ws, cursor, sucursales, dias_t, dias_mes)
    cursor += 2
    cursor = _block_reporte_diario(ws, cursor, sucursales, dias_t, dias_mes)
    cursor += 2
    cursor = _block_semaforo(ws, cursor)
    cursor += 2
    _block_seguimiento_diario(ws, cursor, sucursales, dias_mes)

    # Anchos sugeridos (col A grande para etiquetas).
    ws.column_dimensions["A"].width = 32
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18


# ── Estilos compartidos para los bloques ─────────────────────────────────
_BLOCK_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_BLOCK_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")


def _block_title(ws: Worksheet, row: int, text: str, ncols: int) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _BLOCK_HEADER_FONT
    cell.fill = _BLOCK_HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = _BLOCK_HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1


def _write_subheader(ws: Worksheet, row: int, headers: list[str]) -> None:
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font = Font(bold=True, color="1F4E78")
        cell.fill = _SUBHEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN


def _block_desglose_metas(
    ws: Worksheet, start_row: int, sucursales: list[dict], dias_t: int, dias_mes: int
) -> int:
    """Bloque 1: tabla con Meta Mensual / Diaria / Acumulada ayer / Acumulada hoy."""
    headers = [
        "Sucursal",
        "Meta Mensual (S/)",
        f"Meta Diaria (÷{dias_mes})",
        f"Meta Acumulada al día {max(0, dias_t - 1)}",
        f"Meta Acumulada al día {dias_t}",
    ]
    row = _block_title(ws, start_row, "🎯 Desglose de metas diarias", len(headers))
    _write_subheader(ws, row, headers)
    row += 1
    total_meta = 0.0
    total_meta_diaria = 0.0
    total_meta_ayer = 0.0
    total_meta_hoy = 0.0
    for s in sucursales:
        meta = _num(s.get("meta_mensual"))
        md = _num(s.get("meta_diaria"))
        meta_ayer = md * max(0, dias_t - 1) if md is not None else None
        meta_hoy = md * dias_t if md is not None else None
        ws.cell(row=row, column=1, value=s.get("nombre")).border = BORDER_THIN
        for c_idx, v, fmt in [
            (2, meta, MONEY_FMT),
            (3, md, MONEY_FMT),
            (4, meta_ayer, MONEY_FMT),
            (5, meta_hoy, MONEY_FMT),
        ]:
            cell = ws.cell(row=row, column=c_idx, value=v)
            cell.number_format = fmt
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="right")
        if meta is not None: total_meta += meta
        if md is not None: total_meta_diaria += md
        if meta_ayer is not None: total_meta_ayer += meta_ayer
        if meta_hoy is not None: total_meta_hoy += meta_hoy
        row += 1
    # Fila TOTAL
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = BORDER_THIN
    ws.cell(row=row, column=1).fill = _SUBHEADER_FILL
    for c_idx, v in [(2, total_meta), (3, total_meta_diaria), (4, total_meta_ayer), (5, total_meta_hoy)]:
        cell = ws.cell(row=row, column=c_idx, value=v)
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True)
        cell.fill = _SUBHEADER_FILL
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="right")
    return row


def _block_reporte_diario(
    ws: Worksheet, start_row: int, sucursales: list[dict], dias_t: int, dias_mes: int
) -> int:
    """Bloque 2: tabla pivot — Indicador en filas, Sucursal en columnas (+ Total)."""
    suc_names = [s.get("nombre") or "?" for s in sucursales]
    headers = ["Indicador", "Fórmula", *suc_names, "Total"]
    row = _block_title(ws, start_row, f"📊 Reporte diario (corte al día {dias_t} de {dias_mes})", len(headers))
    _write_subheader(ws, row, headers)
    row += 1

    def _suc_val(suc: dict, key: str) -> float | None:
        return _num(suc.get(key))

    def _total_of(key: str) -> float | None:
        vals = [_suc_val(s, key) for s in sucursales]
        vals = [v for v in vals if v is not None]
        return sum(vals) if vals else None

    indicadores = [
        ("Venta Real (1–{}/mes)".format(dias_t), "Suma diaria del mes", "venta_acumulada", MONEY_FMT),
        ("Gap vs Meta al 50%",                   "Venta Real − Meta ÷ 2", "gap_vs_50pct",   MONEY_FMT),
        ("% Avance del mes",                     "Venta Real ÷ Meta Mensual", "avance_pct", PCT_FMT),
        ("Ritmo actual diario",                  f"Venta Real ÷ {dias_t}", "ritmo_actual",  MONEY_FMT),
        (f"Ritmo requerido ({dias_t+1}–{dias_mes})", f"(Meta − Venta) ÷ {max(0, dias_mes - dias_t)}", "ritmo_requerido", MONEY_FMT),
        ("Proyección mensual",                   f"Ritmo actual × {dias_mes}", "proyeccion", MONEY_FMT),
    ]
    for label, formula, key, fmt in indicadores:
        # Etiqueta + fórmula
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = BORDER_THIN
        f_cell = ws.cell(row=row, column=2, value=formula)
        f_cell.font = Font(italic=True, color="808080", size=10)
        f_cell.border = BORDER_THIN
        # Valor por sucursal
        for c_idx, s in enumerate(sucursales, start=3):
            v = _suc_val(s, key)
            cell = ws.cell(row=row, column=c_idx, value=v)
            cell.number_format = fmt
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="right")
            # Coloreo gap y ritmo requerido
            if key == "gap_vs_50pct" and isinstance(v, (int, float)):
                cell.fill = FILL_OK if v >= 0 else FILL_BAD
                cell.font = Font(bold=True)
            if key == "ritmo_requerido":
                ra = _suc_val(s, "ritmo_actual")
                if isinstance(v, (int, float)) and isinstance(ra, (int, float)) and ra > 0:
                    cell.fill = FILL_BAD if v > ra * 1.15 else FILL_OK
                    cell.font = Font(bold=True)
        # Total
        total = _total_of(key) if key != "avance_pct" else None
        if key == "avance_pct":  # avance promedio no es sumable, calcular sobre totales reales
            mt = _total_of("meta_mensual")
            vt = _total_of("venta_acumulada")
            total = (vt / mt * 100) if (mt and vt is not None) else None
        elif key == "ritmo_actual" and dias_t:
            vt = _total_of("venta_acumulada")
            total = (vt / dias_t) if vt is not None else None
        elif key == "ritmo_requerido" and (dias_mes - dias_t) > 0:
            mt = _total_of("meta_mensual")
            vt = _total_of("venta_acumulada")
            total = ((mt - vt) / (dias_mes - dias_t)) if (mt and vt is not None) else None
        elif key == "proyeccion" and dias_t:
            vt = _total_of("venta_acumulada")
            total = (vt / dias_t * dias_mes) if vt is not None else None
        cell = ws.cell(row=row, column=3 + len(sucursales), value=total)
        cell.number_format = fmt
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)
        cell.fill = _SUBHEADER_FILL
        row += 1
    return row - 1


def _block_semaforo(ws: Worksheet, start_row: int) -> int:
    """Bloque 3: tabla con las reglas del semáforo y la acción esperada."""
    headers = ["Estado", "Regla", "Acción"]
    row = _block_title(ws, start_row, "✅ Semáforo del avance vs meta", len(headers))
    _write_subheader(ws, row, headers)
    row += 1
    reglas = [
        ("🟢 A tiempo",       "Venta Real ≥ Meta Acumulada al día",                "Mantener ritmo",        FILL_OK),
        ("🟡 Desfasado leve", "Venta Real entre 80% y 100% de Meta Acumulada",     "Acelerar promos",       PatternFill("solid", fgColor="FFF2A8")),
        ("🔴 En riesgo",      "Venta Real < 80% de Meta Acumulada",                 "Revisar quiebres y activar ofertas urgentes", FILL_BAD),
    ]
    for estado, regla, accion, fill in reglas:
        ws.cell(row=row, column=1, value=estado).border = BORDER_THIN
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=regla).border = BORDER_THIN
        ws.cell(row=row, column=3, value=accion).border = BORDER_THIN
        ws.cell(row=row, column=3).font = Font(italic=True)
        row += 1
    return row - 1


def _block_seguimiento_diario(
    ws: Worksheet, start_row: int, sucursales: list[dict], dias_mes: int
) -> int:
    """Bloque 4: una fila por cada día del mes con meta · venta · gap · acumulados.

    Si hay varias sucursales, usamos los valores totalizados (sumando por día).
    Días futuros: vacíos en venta/gap/acum, pero con meta diaria/acumulada visible.
    """
    headers = [
        "Fecha", "Día", "Meta Diaria (S/)", "Venta Real (S/)", "Diferencia",
        "Acum. Esperado", "Acum. Real", "Gap Acumulado",
    ]
    row = _block_title(ws, start_row, "📋 Tabla diaria de seguimiento (mes completo)", len(headers))
    _write_subheader(ws, row, headers)
    row += 1

    # Combinar series de todas las sucursales por fecha
    series = [s.get("serie") or [] for s in sucursales]
    if not series:
        return row
    n = len(series[0])
    meta_diaria_total = sum((_num(s.get("meta_diaria")) or 0) for s in sucursales)

    acum_real = 0.0
    acum_esperado = 0.0
    for i in range(n):
        # Tomamos la fecha y estado del primer elemento (es la misma para todas las sucs).
        fecha = series[0][i].get("fecha")
        dia = series[0][i].get("dia")
        estado = series[0][i].get("estado")
        venta_dia_total = sum(
            (_num(s[i].get("venta_dia")) or 0) for s in series if i < len(s)
        )
        if estado == "futuro":
            venta_show = None
            diff = None
            acum_real_show = None
            gap_acum = None
        else:
            acum_real += venta_dia_total
            venta_show = round(venta_dia_total, 2)
            diff = round(venta_dia_total - meta_diaria_total, 2)
            acum_real_show = round(acum_real, 2)
        acum_esperado += meta_diaria_total
        gap_acum = (round(acum_real_show - acum_esperado, 2)
                    if acum_real_show is not None else None)
        vals = [
            fecha,
            dia,
            round(meta_diaria_total, 2) if meta_diaria_total else None,
            venta_show,
            diff,
            round(acum_esperado, 2),
            acum_real_show,
            gap_acum,
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c_idx, value=v)
            cell.border = BORDER_THIN
            if c_idx in (3, 4, 5, 6, 7, 8):
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 2:
                cell.alignment = Alignment(horizontal="center")
            if estado == "futuro":
                cell.fill = FILL_FUTURO
            elif estado == "hoy":
                cell.fill = _SUBHEADER_FILL
                cell.font = Font(bold=True)
        # Color del Gap Acumulado
        gap_cell = ws.cell(row=row, column=8)
        if isinstance(gap_acum, (int, float)):
            gap_cell.fill = FILL_OK if gap_acum >= 0 else FILL_BAD
            gap_cell.font = Font(bold=True)
        row += 1
    return row - 1


def build_daily_workbook(daily: dict, *, brand_name: str = "KAWII") -> Workbook:
    """Construye el workbook del informe diario con sus 3 pestañas."""
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_ticket(wb, daily)
    _sheet_transacciones(wb, daily)
    _sheet_meta_diaria(wb, daily)
    return wb
