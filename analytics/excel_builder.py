"""
Builder de reportes Excel para los módulos KAWII Matrix (04, 05, ...).

Layout:
    1) Portada       — metadata del reporte
    2) Índice        — listado de departamentos con conteos y links
    3) Resumen       — distribución por clasificación (global)
    4..N) Una hoja por DEPARTAMENTO, con filas agrupadas por
          Categoría → Subcategoría usando outline colapsable
          (botones +/- de Excel). Cada hoja tiene autofilter completo.

Diseñado para ser reutilizado: los scripts `generar_reporte_*.py`
solo arman los metadata del módulo y delegan acá la construcción.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ──────────────────────────────────────────────────────────────────────────
# Estilos compartidos
# ──────────────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Filas de agrupamiento (categoría / subcategoría dentro de la hoja-departamento)
CAT_FILL = PatternFill("solid", fgColor="2E75B6")     # azul medio
CAT_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBCAT_FILL = PatternFill("solid", fgColor="BDD7EE")  # azul claro
SUBCAT_FONT = Font(bold=True, color="1F4E78", size=10)

# Colores rotativos para las pestañas de las hojas por departamento.
# Le da un look más profesional y se distinguen a simple vista.
TAB_COLORS = [
    "1F4E78", "2E75B6", "548235", "BF8F00", "C00000",
    "7030A0", "385723", "C65911", "203864", "843C0C",
]

# Paleta de colores para diagnósticos (Clasificación KAWII).
# Cada tupla = (substring que matchea, fill RGB, font RGB).
# El orden importa: el primer match gana.
COLOR_DIAGNOSTICOS: list[tuple[str, str, str]] = [
    ("🚨", "FFB3B3", "8B0000"), ("URGENTE", "FFB3B3", "8B0000"),
    ("QUIEBRE", "FFB3B3", "8B0000"), ("TRANSFERIR", "FFD699", "8B4500"),
    ("REBALANCEAR", "FFD699", "8B4500"), ("COMPRAR YA", "FFB3B3", "8B0000"),
    ("CORRUPT", "FF6B6B", "FFFFFF"),
    ("🐢", "FFE6B3", "8B5A00"), ("BAJA ROT", "FFE6B3", "8B5A00"),
    ("⚠️", "FFE6B3", "8B5A00"), ("CRÍTICO", "FFE6B3", "8B5A00"),
    ("🥷", "E0BBE4", "4B0082"), ("ESCONDIDO", "E0BBE4", "4B0082"),
    ("🐀", "D4A5A5", "8B0000"), ("PARÁSITO", "D4A5A5", "8B0000"),
    ("💎", "B5EAD7", "006400"), ("JOYA", "B5EAD7", "006400"),
    ("🏆", "B5EAD7", "006400"), ("CAMPEÓN", "B5EAD7", "006400"),
    ("🔥", "FFCCCC", "8B0000"), ("ALTA ROTACIÓN", "FFCCCC", "8B0000"),
    ("🚀", "C7E9C0", "006400"), ("BUILD", "C7E9C0", "006400"),
    ("🌱", "D4F1D4", "006400"), ("NUEVO", "D4F1D4", "006400"),
    ("✅", "D4F1D4", "006400"), ("MANTENER", "D4F1D4", "006400"),
    ("🟢", "D4F1D4", "006400"), ("SANO", "D4F1D4", "006400"),
    ("⚡", "E8F4D4", "4F6B00"), ("MEDIA ROTACIÓN", "E8F4D4", "4F6B00"),
    ("💀", "D3D3D3", "595959"), ("MUERTO", "D3D3D3", "595959"),
    ("⚰️", "D3D3D3", "595959"), ("FRACASO", "D3D3D3", "595959"),
    ("👻", "E8E8E8", "595959"), ("AGOTADO", "E8E8E8", "595959"),
    ("🌀", "F0F0F0", "808080"), ("FANTASMA", "F0F0F0", "808080"),
    ("🧊", "B3D9FF", "003366"), ("EXCESO", "B3D9FF", "003366"),
    ("CAPITAL ESTANCADO", "B3D9FF", "003366"),
    ("🔄", "BFE3F5", "004C66"), ("CICLO CERRADO", "BFE3F5", "004C66"),
    ("🧱", "F5D4A6", "7B3F00"), ("ATASCADO", "F5D4A6", "7B3F00"),
    ("🗑️", "C0C0C0", "595959"), ("DELETE", "C0C0C0", "595959"),
    ("✂️", "FFDAB9", "7B3F00"), ("RACIONALIZAR", "FFDAB9", "7B3F00"),
    ("🔍", "FFF4C7", "7B5C00"), ("INVESTIGAR", "FFF4C7", "7B5C00"),
    ("📊", "E0E7FF", "1B3A8F"), ("OPTIMIZAR", "E0E7FF", "1B3A8F"),
    ("🎯", "FFE0F0", "8B005C"), ("NICHO", "FFE0F0", "8B005C"),
    ("📈", "D4F1D4", "006400"), ("Subiendo", "D4F1D4", "006400"),
    ("📉", "FFE6B3", "8B5A00"), ("Bajando", "FFE6B3", "8B5A00"),
    ("🏪", "F0E68C", "6B5800"), ("EXCLUSIVO", "F0E68C", "6B5800"),
    ("💤", "B3D9FF", "003366"), ("DORMIDO", "B3D9FF", "003366"),
    ("🔀", "FFD699", "8B4500"), ("🆕", "D4F1D4", "006400"),
]

# Columnas que vale la pena agregar (sumar) en las filas-cabecera de
# Categoría/Subcategoría. Si están presentes en `cols`, se suman.
# Las que terminan en "S/" se formatean como moneda.
NUMERIC_AGGREGATABLE = (
    "Unds Vend (90d)",
    "Unds Recib (90d)",
    "Vend Lote Total",
    "Stock Disp",
    "Stock Reserv",
    "Vendido SKU S/",
    "Unds Vend Lifetime",
)


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────


def color_diagnostico(texto: str | None) -> tuple[str | None, str | None]:
    """Devuelve (fill_color, font_color) según la palabra clave del diagnóstico."""
    if not texto:
        return None, None
    s = str(texto)
    for keyword, fill, font in COLOR_DIAGNOSTICOS:
        if keyword in s:
            return fill, font
    return None, None


def ancho_columna(values: Iterable[Any], header: str) -> int:
    """Calcula un ancho razonable para una columna (entre 10 y 50 chars)."""
    max_len = len(str(header))
    for v in list(values)[:200]:
        if v is not None and len(str(v)) > max_len:
            max_len = len(str(v))
    return min(max(max_len + 2, 10), 50)


_INVALID_SHEET_CHARS = set('[]:*?/\\')


def _sheet_name(base: str, used: set[str]) -> str:
    """Sanitiza un nombre de hoja: max 31 chars, sin caracteres prohibidos, único."""
    cleaned = "".join(c for c in (base or "Sin nombre") if c not in _INVALID_SHEET_CHARS).strip()
    cleaned = cleaned[:31] or "Hoja"
    if cleaned not in used:
        used.add(cleaned)
        return cleaned
    # Resolver colisión: sufijo numérico que respete el límite de 31 chars
    for i in range(2, 100):
        suffix = f" ({i})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise ValueError(f"No se pudo generar nombre único para '{base}'")


def _coerce_numeric(val: Any) -> float | int:
    """Convierte a número para sumas. Si no se puede, devuelve 0."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0


def _is_money_col(name: str) -> bool:
    """Detecta si una columna representa dinero (para formatearla como S/)."""
    n = name.strip()
    return n.endswith("S/") or "S/" in n


# ──────────────────────────────────────────────────────────────────────────
# Construcción del workbook
# ──────────────────────────────────────────────────────────────────────────


def build_workbook(
    *,
    cols: list[str],
    rows: list[tuple],
    modulo_id: str,
    titulo: str,
    sql_file: str,
    descripcion: str,
    classification_col: str | None,
    elapsed_seconds: float,
    brand_name: str,
) -> Workbook:
    """
    Construye el Workbook completo.

    `classification_col` es el nombre de la columna que contiene la
    Clasificación KAWII (puede haberse renombrado vía settings).
    """
    wb = Workbook()
    wb.remove(wb.active)  # arranco limpio, agrego las hojas yo

    # Resolver índices de columnas clave una sola vez
    idx = {name: cols.index(name) for name in cols}
    idx_sucursal = idx.get("Sucursal")
    idx_depto = idx.get("Departamento")
    idx_cat = idx.get("Categoría")
    idx_subcat = idx.get("Subcategoría")
    idx_diag = idx.get(classification_col) if classification_col else None
    idx_aggregable = [(cols.index(c), c) for c in NUMERIC_AGGREGATABLE if c in idx]

    # Agrupar filas por departamento
    by_dept: dict[str, list[tuple]] = defaultdict(list)
    if idx_depto is not None:
        for r in rows:
            dept = str(r[idx_depto]) if r[idx_depto] else "— Sin departamento"
            by_dept[dept].append(r)
    else:
        # SQL sin columna Departamento → todo va a una sola hoja
        by_dept["Datos"] = list(rows)

    # 1) Portada
    _build_portada(wb, brand_name, modulo_id, titulo, sql_file, descripcion,
                   len(rows), elapsed_seconds)

    # 2) Índice — listado de departamentos con conteos y links
    _build_indice(wb, by_dept, cols, idx_aggregable, idx_diag)

    # 3) Resumen global (distribución por clasificación)
    if classification_col and idx_diag is not None and rows:
        _build_resumen(wb, rows, idx_diag, classification_col)

    # 4..N) Una hoja por departamento, con outline colapsable
    used_sheet_names = set(wb.sheetnames)
    for tab_idx, (dept_name, dept_rows) in enumerate(
        sorted(by_dept.items(), key=lambda kv: -len(kv[1]))
    ):
        _build_hoja_departamento(
            wb,
            dept_name=dept_name,
            dept_rows=dept_rows,
            cols=cols,
            used_sheet_names=used_sheet_names,
            idx_cat=idx_cat,
            idx_subcat=idx_subcat,
            idx_diag=idx_diag,
            idx_aggregable=idx_aggregable,
            tab_color=TAB_COLORS[tab_idx % len(TAB_COLORS)],
        )

    return wb


# ──────────────────────────────────────────────────────────────────────────
# Hojas individuales
# ──────────────────────────────────────────────────────────────────────────


def _build_portada(
    wb: Workbook, brand_name: str, modulo_id: str, titulo: str,
    sql_file: str, descripcion: str, n_rows: int, elapsed: float,
) -> None:
    ws = wb.create_sheet("Portada")
    ws.sheet_properties.tabColor = "1F4E78"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws["A1"] = f"REPORTE {brand_name.upper()} PLUS"
    ws["A1"].font = Font(size=24, bold=True, color="1F4E78")
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")
    metadata = [
        ("Módulo:", f"{modulo_id} — {titulo}"),
        ("Generado:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("SQL ejecutado:", sql_file),
        ("Filas devueltas:", n_rows),
        ("Tiempo ejecución:", f"{elapsed:.2f} segundos"),
    ]
    for i, (k, v) in enumerate(metadata, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws["A9"] = "Descripción:"
    ws["A9"].font = Font(bold=True)
    ws["B9"] = descripcion
    ws["B9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[9].height = 80


def _build_indice(
    wb: Workbook,
    by_dept: dict[str, list[tuple]],
    cols: list[str],
    idx_aggregable: list[tuple[int, str]],
    idx_diag: int | None,
) -> None:
    """Hoja índice con un link a cada hoja de departamento."""
    ws = wb.create_sheet("Índice")
    ws.sheet_properties.tabColor = "548235"
    ws["A1"] = "Índice de departamentos"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws.merge_cells("A1:F1")

    headers = ["Departamento", "SKUs"] + [name for _, name in idx_aggregable]
    if idx_diag is not None:
        headers.append("Urgentes")  # SKUs con "🚨" / "URGENTE"
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
    ws.row_dimensions[3].height = 28

    # Vamos a ordenar los departamentos por cantidad de SKUs descendente
    sorted_depts = sorted(by_dept.items(), key=lambda kv: -len(kv[1]))
    used = set(wb.sheetnames)
    # NOTA: el nombre de hoja se asigna ABAJO en _build_hoja_departamento.
    # Aquí precomputamos los nombres en el mismo orden para que los links
    # apunten a la hoja correcta.
    dept_to_sheet = {d: _sheet_name(d, used) for d, _ in sorted_depts}

    for r_offset, (dept, dept_rows) in enumerate(sorted_depts):
        r_idx = 4 + r_offset
        # Link al nombre del departamento
        link_cell = ws.cell(row=r_idx, column=1, value=dept)
        link_cell.hyperlink = f"#'{dept_to_sheet[dept]}'!A1"
        link_cell.font = Font(color="1F4E78", underline="single", bold=True)
        link_cell.border = BORDER_THIN
        # Cantidad de SKUs
        sku_cell = ws.cell(row=r_idx, column=2, value=len(dept_rows))
        sku_cell.alignment = Alignment(horizontal="right")
        sku_cell.border = BORDER_THIN
        # Sumas agregables
        for col_offset, (col_idx, col_name) in enumerate(idx_aggregable, start=3):
            total = sum(_coerce_numeric(r[col_idx]) for r in dept_rows)
            cell = ws.cell(row=r_idx, column=col_offset, value=round(total, 2))
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER_THIN
            if _is_money_col(col_name):
                cell.number_format = '"S/ "#,##0.00'
            else:
                cell.number_format = "#,##0.##"
        # Urgentes
        if idx_diag is not None:
            urgentes = sum(
                1 for r in dept_rows
                if r[idx_diag] and ("🚨" in str(r[idx_diag]) or "URGENTE" in str(r[idx_diag]).upper())
            )
            cell = ws.cell(row=r_idx, column=len(headers), value=urgentes)
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER_THIN
            if urgentes > 0:
                cell.fill = PatternFill("solid", fgColor="FFB3B3")
                cell.font = Font(bold=True, color="8B0000")

    # Fila de totales
    tr = 4 + len(sorted_depts)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=sum(len(r) for r in by_dept.values())).font = Font(bold=True)
    ws.cell(row=tr, column=2).alignment = Alignment(horizontal="right")
    for col_offset, (col_idx, col_name) in enumerate(idx_aggregable, start=3):
        total = sum(_coerce_numeric(r[col_idx]) for rows in by_dept.values() for r in rows)
        cell = ws.cell(row=tr, column=col_offset, value=round(total, 2))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        if _is_money_col(col_name):
            cell.number_format = '"S/ "#,##0.00'

    # Anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    for col_offset in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_offset)].width = 16
    ws.freeze_panes = "A4"

    # Devolver el mapa para que _build_hoja_departamento use los mismos nombres
    wb._dept_to_sheet = dept_to_sheet  # type: ignore[attr-defined]


def _build_resumen(
    wb: Workbook, rows: list[tuple], idx_diag: int, classification_col: str,
) -> None:
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = "7030A0"
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws["A1"] = f'Distribución por "{classification_col}"'
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:C1")
    for c_idx, hdr in enumerate(["Diagnóstico", "Cantidad", "%"], 1):
        c = ws.cell(row=3, column=c_idx, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER_THIN
    cnt = Counter(str(r[idx_diag]) if r[idx_diag] else "(sin)" for r in rows)
    total = sum(cnt.values())
    for r_idx, (diag, count) in enumerate(sorted(cnt.items(), key=lambda x: -x[1]), 4):
        ws.cell(row=r_idx, column=1, value=diag).border = BORDER_THIN
        ws.cell(row=r_idx, column=2, value=count).border = BORDER_THIN
        ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=r_idx, column=3, value=f"{100 * count / total:.1f}%").border = BORDER_THIN
        ws.cell(row=r_idx, column=3).alignment = Alignment(horizontal="right")
        fc, fnt = color_diagnostico(diag)
        if fc:
            ws.cell(row=r_idx, column=1).fill = PatternFill("solid", fgColor=fc)
            ws.cell(row=r_idx, column=1).font = Font(color=fnt, bold=True)
    tr = 4 + len(cnt)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=2, value=total).font = Font(bold=True)
    ws.cell(row=tr, column=3, value="100%").font = Font(bold=True)


def _build_hoja_departamento(
    wb: Workbook,
    *,
    dept_name: str,
    dept_rows: list[tuple],
    cols: list[str],
    used_sheet_names: set[str],
    idx_cat: int | None,
    idx_subcat: int | None,
    idx_diag: int | None,
    idx_aggregable: list[tuple[int, str]],
    tab_color: str,
) -> None:
    """Crea una hoja con los SKUs de UN departamento, agrupados por Cat→Subcat."""
    # Reutilizar el nombre precomputado en el Índice si existe
    dept_to_sheet = getattr(wb, "_dept_to_sheet", None)
    if dept_to_sheet and dept_name in dept_to_sheet:
        sheet_name = dept_to_sheet[dept_name]
        used_sheet_names.add(sheet_name)
    else:
        sheet_name = _sheet_name(dept_name, used_sheet_names)

    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    # Botón de outline arriba del grupo (más natural cuando el header
    # de Categoría aparece ANTES de las subcategorías)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Header de columnas
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 32

    # Si no hay columnas Cat/Subcat, escribimos las filas planas y salimos.
    if idx_cat is None or idx_subcat is None:
        _write_data_rows(ws, dept_rows, cols, idx_diag, start_row=2, outline_level=0)
        _finalize_sheet(ws, cols, dept_rows, last_row=1 + len(dept_rows))
        return

    # Agrupar filas: Categoría → Subcategoría → SKUs
    cat_map: dict[str, list[tuple]] = defaultdict(list)
    for r in dept_rows:
        cat_map[str(r[idx_cat]) if r[idx_cat] else "— Sin categoría"].append(r)

    current_row = 2
    # Ordenar categorías por ventas/SKUs desc para que las importantes aparezcan arriba
    for cat in sorted(cat_map.keys(), key=lambda c: -_total_for_sorting(cat_map[c], idx_aggregable)):
        cat_rows = cat_map[cat]
        # Fila-header de Categoría (outline level 0 — siempre visible)
        _write_group_header(
            ws,
            row=current_row,
            level_label=f"📂 {cat}",
            n_skus=len(cat_rows),
            n_cols=len(cols),
            agg_values=_aggregate(cat_rows, idx_aggregable),
            fill=CAT_FILL,
            font=CAT_FONT,
            money_cols=[name for _, name in idx_aggregable if _is_money_col(name)],
            idx_aggregable=idx_aggregable,
        )
        ws.row_dimensions[current_row].outline_level = 0
        current_row += 1

        # Sub-agrupar por subcategoría
        subcat_map: dict[str, list[tuple]] = defaultdict(list)
        for r in cat_rows:
            subcat_map[str(r[idx_subcat]) if r[idx_subcat] else "— Sin subcategoría"].append(r)

        for subcat in sorted(
            subcat_map.keys(), key=lambda s: -_total_for_sorting(subcat_map[s], idx_aggregable)
        ):
            sub_rows = subcat_map[subcat]
            # Fila-header de Subcategoría (outline level 1 — visible cuando expandís cat)
            _write_group_header(
                ws,
                row=current_row,
                level_label=f"    📁 {subcat}",
                n_skus=len(sub_rows),
                n_cols=len(cols),
                agg_values=_aggregate(sub_rows, idx_aggregable),
                fill=SUBCAT_FILL,
                font=SUBCAT_FONT,
                money_cols=[name for _, name in idx_aggregable if _is_money_col(name)],
                idx_aggregable=idx_aggregable,
            )
            ws.row_dimensions[current_row].outline_level = 1
            current_row += 1

            # SKUs (outline level 2 — más profundo, colapsable)
            _write_data_rows(
                ws, sub_rows, cols, idx_diag, start_row=current_row, outline_level=2
            )
            current_row += len(sub_rows)

    _finalize_sheet(ws, cols, dept_rows, last_row=current_row - 1)


# ──────────────────────────────────────────────────────────────────────────
# Helpers de hoja-departamento
# ──────────────────────────────────────────────────────────────────────────


def _aggregate(
    rows: list[tuple], idx_aggregable: list[tuple[int, str]],
) -> dict[int, float]:
    """Suma las columnas agregables para un bloque de filas."""
    return {idx: sum(_coerce_numeric(r[idx]) for r in rows) for idx, _ in idx_aggregable}


def _total_for_sorting(
    rows: list[tuple], idx_aggregable: list[tuple[int, str]],
) -> float:
    """Métrica para ordenar grupos: prefiere ventas S/, si no Unds Vend, si no SKUs."""
    for idx, name in idx_aggregable:
        if _is_money_col(name) and "Vendido SKU" in name:
            return sum(_coerce_numeric(r[idx]) for r in rows)
    for idx, name in idx_aggregable:
        if name == "Unds Vend (90d)":
            return sum(_coerce_numeric(r[idx]) for r in rows)
    return float(len(rows))


def _write_group_header(
    ws: Worksheet,
    *,
    row: int,
    level_label: str,
    n_skus: int,
    n_cols: int,
    agg_values: dict[int, float],
    fill: PatternFill,
    font: Font,
    money_cols: list[str],
    idx_aggregable: list[tuple[int, str]],
) -> None:
    """
    Escribe la fila de cabecera de un grupo (Categoría o Subcategoría):
    - Columna 1: nombre con emoji
    - Columnas agregables: sumas
    - Resto: pintadas del color del grupo (sin valor)
    """
    money_set = set(money_cols)
    # Pintar TODA la fila del color del grupo (para que se vea como header visual)
    for c_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c_idx)
        cell.fill = fill
        cell.font = font
        cell.border = BORDER_THIN
    # Nombre del grupo en columna 1
    ws.cell(row=row, column=1, value=level_label)
    # Conteo de SKUs en columna 2 (suele ser "Departamento" en SQL, sobrescribimos)
    ws.cell(row=row, column=2, value=f"{n_skus} SKU{'s' if n_skus != 1 else ''}")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
    # Valores agregados (sumas) en sus columnas correspondientes
    for col_idx_0based, col_name in idx_aggregable:
        total = agg_values.get(col_idx_0based, 0)
        cell = ws.cell(row=row, column=col_idx_0based + 1, value=round(total, 2))
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="right")
        cell.border = BORDER_THIN
        if col_name in money_set:
            cell.number_format = '"S/ "#,##0.00'
        else:
            cell.number_format = "#,##0.##"


def _write_data_rows(
    ws: Worksheet,
    rows: list[tuple],
    cols: list[str],
    idx_diag: int | None,
    *,
    start_row: int,
    outline_level: int,
) -> None:
    """Escribe filas de datos (SKUs) con borde, color de diagnóstico y outline."""
    for offset, row in enumerate(rows):
        r_idx = start_row + offset
        diag_val = (
            str(row[idx_diag]) if idx_diag is not None and row[idx_diag] else None
        )
        diag_fill, diag_font = color_diagnostico(diag_val) if diag_val else (None, None)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER_THIN
            if idx_diag is not None and c_idx - 1 == idx_diag and diag_fill:
                cell.fill = PatternFill("solid", fgColor=diag_fill)
                cell.font = Font(color=diag_font, bold=True)
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        if outline_level > 0:
            ws.row_dimensions[r_idx].outline_level = outline_level


def _finalize_sheet(
    ws: Worksheet, cols: list[str], all_rows: list[tuple], last_row: int,
) -> None:
    """Aplica anchos, freeze panes y autofilter al rango completo de datos."""
    for c_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(c_idx)
        values = [r[c_idx - 1] for r in all_rows[:200]]
        ws.column_dimensions[col_letter].width = ancho_columna(values, col_name)
    ws.freeze_panes = "A2"
    if last_row >= 2:
        last_col_letter = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
