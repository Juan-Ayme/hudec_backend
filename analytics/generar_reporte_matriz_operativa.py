#!/usr/bin/env python3
"""
Generador de reporte Excel — Módulo 05: Matriz Operativa (90d + histórico)
=====================================================================
Lee el SQL 'matriz_kawii_v2.sql' y produce un Excel con portada,
datos coloreados por clasificación y resumen.

Uso: python generar_reporte.py
Salida: ../reportes_excel/[fecha]/05_Matriz_Operativa.xlsx
"""
import sys
import psycopg2
from datetime import datetime
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import re
import os
# Asegurar path para importación correcta de harvester y app
sys.path.append(str(Path(__file__).resolve().parent.parent))
from harvester.config import (
    DB_CONFIG,
    OFFICES_TIENDA,
    TIPOS_VENTA,
    TIPOS_DEVOLUCION,
    TIPOS_TRASLADO,
    EXCLUDED_DEPARTMENTS,
    EXCLUDED_CATEGORIES,
)
from app.config import get_settings

sys.stdout.reconfigure(encoding='utf-8')

# ════ CONFIG DEL MÓDULO ════
settings = get_settings()

MODULO_ID = '05'
TITULO = 'Matriz Operativa (90d + histórico)'
# Nombre del SQL (solo para mostrar en la portada del Excel)
SQL_FILE = '05_matriz_operativa.sql'
# Path relativo desde analytics/ al SQL del backend (única fuente de verdad)
SQL_PATH = '../app/kawii_matrix/sql/05_matriz_operativa.sql'
DESCRIPCION = ('Vista operativa principal por SKU. Combina métricas de 90 días '
               'con contexto histórico (lifetime, tendencia 90d, mejor mes, '
               'sell-through lifetime). Velocidad calculada del LOTE ACTUAL '
               '(desde última recepción, con piso de 7 días).')
COL_DIAGNOSTICO = settings.CLASSIFICATION_LABEL
COL_NIVEL = None

# ════ ESTILOS Y PALETAS (idénticos en todos los módulos) ════
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER_THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
COLOR_DIAGNOSTICOS = [
    ('🚨', 'FFB3B3', '8B0000'), ('URGENTE', 'FFB3B3', '8B0000'),
    ('QUIEBRE', 'FFB3B3', '8B0000'), ('TRANSFERIR', 'FFD699', '8B4500'),
    ('REBALANCEAR', 'FFD699', '8B4500'), ('COMPRAR YA', 'FFB3B3', '8B0000'),
    ('CORRUPT', 'FF6B6B', 'FFFFFF'),
    ('🐢', 'FFE6B3', '8B5A00'), ('BAJA ROT', 'FFE6B3', '8B5A00'),
    ('⚠️', 'FFE6B3', '8B5A00'), ('CRÍTICO', 'FFE6B3', '8B5A00'),
    ('🥷', 'E0BBE4', '4B0082'), ('ESCONDIDO', 'E0BBE4', '4B0082'),
    ('🐀', 'D4A5A5', '8B0000'), ('PARÁSITO', 'D4A5A5', '8B0000'),
    ('💎', 'B5EAD7', '006400'), ('JOYA', 'B5EAD7', '006400'),
    ('🏆', 'B5EAD7', '006400'), ('CAMPEÓN', 'B5EAD7', '006400'),
    ('🔥', 'FFCCCC', '8B0000'), ('ALTA ROTACIÓN', 'FFCCCC', '8B0000'),
    ('🚀', 'C7E9C0', '006400'), ('BUILD', 'C7E9C0', '006400'),
    ('🌱', 'D4F1D4', '006400'), ('NUEVO', 'D4F1D4', '006400'),
    ('✅', 'D4F1D4', '006400'), ('MANTENER', 'D4F1D4', '006400'),
    ('🟢', 'D4F1D4', '006400'), ('SANO', 'D4F1D4', '006400'),
    ('⚡', 'E8F4D4', '4F6B00'), ('MEDIA ROTACIÓN', 'E8F4D4', '4F6B00'),
    ('💀', 'D3D3D3', '595959'), ('MUERTO', 'D3D3D3', '595959'),
    ('⚰️', 'D3D3D3', '595959'), ('FRACASO', 'D3D3D3', '595959'),
    ('👻', 'E8E8E8', '595959'), ('AGOTADO', 'E8E8E8', '595959'),
    ('🌀', 'F0F0F0', '808080'), ('FANTASMA', 'F0F0F0', '808080'),
    ('🧊', 'B3D9FF', '003366'), ('EXCESO', 'B3D9FF', '003366'),
    ('CAPITAL ESTANCADO', 'B3D9FF', '003366'),
    ('🔄', 'BFE3F5', '004C66'), ('CICLO CERRADO', 'BFE3F5', '004C66'),
    ('🧱', 'F5D4A6', '7B3F00'), ('ATASCADO', 'F5D4A6', '7B3F00'),
    ('🗑️', 'C0C0C0', '595959'), ('DELETE', 'C0C0C0', '595959'),
    ('✂️', 'FFDAB9', '7B3F00'), ('RACIONALIZAR', 'FFDAB9', '7B3F00'),
    ('🔍', 'FFF4C7', '7B5C00'), ('INVESTIGAR', 'FFF4C7', '7B5C00'),
    ('📊', 'E0E7FF', '1B3A8F'), ('OPTIMIZAR', 'E0E7FF', '1B3A8F'),
    ('🎯', 'FFE0F0', '8B005C'), ('NICHO', 'FFE0F0', '8B005C'),
    ('📈', 'D4F1D4', '006400'), ('Subiendo', 'D4F1D4', '006400'),
    ('📉', 'FFE6B3', '8B5A00'), ('Bajando', 'FFE6B3', '8B5A00'),
    ('🏪', 'F0E68C', '6B5800'), ('EXCLUSIVO', 'F0E68C', '6B5800'),
    ('💤', 'B3D9FF', '003366'), ('DORMIDO', 'B3D9FF', '003366'),
    ('🔀', 'FFD699', '8B4500'), ('🆕', 'D4F1D4', '006400'),
]
COLOR_NIVELES = {
    'DEPARTAMENTO':  ('1F4E78', 'FFFFFF', True),
    'CATEGORÍA':     ('2E75B6', 'FFFFFF', True),
    'SUBCATEGORÍA':  ('BDD7EE', '1F4E78', True),
    'SKU':           ('FFFFFF', '000000', False),
}


def color_diagnostico(texto):
    if not texto:
        return None, None
    s = str(texto)
    for keyword, fill, font in COLOR_DIAGNOSTICOS:
        if keyword in s:
            return fill, font
    return None, None


def ancho_columna(values, header):
    max_len = len(str(header))
    for v in values[:200]:
        if v is not None and len(str(v)) > max_len:
            max_len = len(str(v))
    return min(max(max_len + 2, 10), 50)


def generar(sql_path: Path, output_path: Path):
    sql = sql_path.read_text(encoding='utf-8')
    settings = get_settings()
    params = {
        "sucursales_objetivo": list(OFFICES_TIENDA),
        "tipos_venta": list(TIPOS_VENTA),
        "tipos_devolucion": list(TIPOS_DEVOLUCION),
        "tipos_traslado": list(TIPOS_TRASLADO),
        "excluded_departments": list(EXCLUDED_DEPARTMENTS),
        "excluded_categories": list(EXCLUDED_CATEGORIES),
        "timezone": settings.TIMEZONE,
    }
    sql_escaped = sql.replace('%', '%%')
    sql_converted = re.sub(r'(?<!:):([a-zA-Z0-9_]+)\b', r'%(\1)s', sql_escaped)
    with psycopg2.connect(**DB_CONFIG) as conn:
        cur = conn.cursor()
        t0 = datetime.now()
        cur.execute(sql_converted, params)
        cols = [settings.CLASSIFICATION_LABEL if d[0] == "Clasificación" else d[0] for d in cur.description]
        rows = cur.fetchall()
        elapsed = (datetime.now() - t0).total_seconds()
    print(f"  → SQL ejecutado: {len(rows)} filas en {elapsed:.2f}s")

    wb = Workbook()

    # Portada
    ws = wb.active
    ws.title = 'Portada'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    ws['A1'] = f'REPORTE {settings.BRAND_NAME.upper()} PLUS'
    ws['A1'].font = Font(size=24, bold=True, color='1F4E78')
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')
    metadata = [
        ('Módulo:', f"{MODULO_ID} — {TITULO}"),
        ('Generado:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
        ('SQL ejecutado:', SQL_FILE),
        ('Filas devueltas:', len(rows)),
        ('Tiempo ejecución:', f"{elapsed:.2f} segundos"),
    ]
    for i, (k, v) in enumerate(metadata, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws['A9'] = 'Descripción:'
    ws['A9'].font = Font(bold=True)
    ws['B9'] = DESCRIPCION
    ws['B9'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[9].height = 80

    # Datos
    ws = wb.create_sheet('Datos')
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
    ws.row_dimensions[1].height = 32

    idx_diag = cols.index(COL_DIAGNOSTICO) if COL_DIAGNOSTICO and COL_DIAGNOSTICO in cols else None
    idx_nivel = cols.index(COL_NIVEL) if COL_NIVEL and COL_NIVEL in cols else None

    for r_idx, row in enumerate(rows, 2):
        nivel_value = str(row[idx_nivel]) if idx_nivel is not None and row[idx_nivel] else None
        nivel_color = COLOR_NIVELES.get(nivel_value) if nivel_value else None
        diag_value = str(row[idx_diag]) if idx_diag is not None and row[idx_diag] else None
        diag_fill, diag_font = color_diagnostico(diag_value) if diag_value else (None, None)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER_THIN
            if nivel_color:
                fc, fnt, bold = nivel_color
                if fc != 'FFFFFF':
                    cell.fill = PatternFill('solid', fgColor=fc)
                cell.font = Font(color=fnt, bold=bold)
            if c_idx - 1 == idx_diag and diag_fill:
                cell.fill = PatternFill('solid', fgColor=diag_fill)
                cell.font = Font(color=diag_font, bold=True)
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    for c_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(c_idx)
        values = [row[c_idx-1] for row in rows]
        ws.column_dimensions[col_letter].width = ancho_columna(values, col_name)
    ws.freeze_panes = 'A2'
    if rows:
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f'A1:{last_col}{len(rows)+1}'

    # Resumen
    if COL_DIAGNOSTICO and idx_diag is not None and rows:
        ws = wb.create_sheet('Resumen')
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws['A1'] = f'Distribución por "{COL_DIAGNOSTICO}"'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        for c_idx, hdr in enumerate(['Diagnóstico', 'Cantidad', '%'], 1):
            c = ws.cell(row=3, column=c_idx, value=hdr)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = HEADER_ALIGN
            c.border = BORDER_THIN
        cnt = Counter(str(r[idx_diag]) if r[idx_diag] else '(sin)' for r in rows)
        total = sum(cnt.values())
        for r_idx, (diag, count) in enumerate(sorted(cnt.items(), key=lambda x: -x[1]), 4):
            ws.cell(row=r_idx, column=1, value=diag).border = BORDER_THIN
            ws.cell(row=r_idx, column=2, value=count).border = BORDER_THIN
            ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal='right')
            ws.cell(row=r_idx, column=3, value=f'{100*count/total:.1f}%').border = BORDER_THIN
            ws.cell(row=r_idx, column=3).alignment = Alignment(horizontal='right')
            fc, fnt = color_diagnostico(diag)
            if fc:
                ws.cell(row=r_idx, column=1).fill = PatternFill('solid', fgColor=fc)
                ws.cell(row=r_idx, column=1).font = Font(color=fnt, bold=True)
        tr = 4 + len(cnt)
        ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=tr, column=2, value=total).font = Font(bold=True)
        ws.cell(row=tr, column=3, value='100%').font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    here = Path(__file__).parent
    sql_path = (here / SQL_PATH).resolve()
    if not sql_path.exists():
        print(f"✗ SQL no encontrado: {sql_path}")
        sys.exit(1)
    fecha = datetime.now().strftime('%Y-%m-%d')
    safe_title = TITULO.replace('—', '-').replace('/', '_').replace(':', '').replace(' ', '_')
    filename = f"{MODULO_ID}_{safe_title}.xlsx"
    # Output en analytics/reportes_excel/ (dentro del backend)
    output_path = here / 'reportes_excel' / fecha / filename
    print(f"\n[{MODULO_ID}] {TITULO}")
    try:
        generar(sql_path, output_path)
        print(f"  ✓ Generado: {output_path}")
    except Exception as e:
        print(f"  ✗ ERROR: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
