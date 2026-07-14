"""Builder Excel del **Informe Compras & Catálogo**.

Reporte enfocado en lo accionable HOY:
    🎯 Memo Ejecutivo — primera hoja para el gerente: 4 KPIs hero, Top 15 SKUs
       priorizados por venta perdida estimada, breakdown por sucursal y por
       departamento, footer con call-to-action a las pestañas de detalle.
    🏬 Hojas por Departamento — jerarquía DEPT→CAT→SUBCAT→SKU del builder
       ejecutivo (filtrada a quiebres reales).
    🏷 Venta por categoría — ranking + ticket promedio.

Consume:
    - `filtered_rows`: filas de la matriz 04b ya filtradas a quiebre real
      (severidades 🔴 Crítico y 🟠 Alta).
    - `venta_categoria`: ranking de categorías del weekly_board.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analytics.excel_builder import (
    BORDER_THIN,
    HEADER_ALIGN,
    HEADER_FILL,
    HEADER_FONT,
    ancho_columna,
)
from analytics.excel_executive import (
    C_TITLE_BG, C_TITLE_FG, C_META_BG, C_META_FG,
    C_TOTAL_BG, C_TOTAL_LBL, C_TOTAL_FG, C_NOTE_FG,
    C_DEPT_BG, C_DEPT_FG,
    C_CAT_BG, C_CAT_FG,
    C_SUBCAT_BG, C_SUBCAT_FG,
)

MONEY_FMT = '"S/ "#,##0.00'
PCT_FMT = '0.0"%"'
INT_FMT = "#,##0"
TITLE_FONT = Font(size=16, bold=True, color="1F4E78")

TAB_QUIEBRE = "C00000"
TAB_CAT = "548235"

_SEV_FILLS: dict[str, PatternFill] = {
    "🔴 Crítico": PatternFill("solid", fgColor="FFB3B3"),
    "🟠 Alta":    PatternFill("solid", fgColor="FFD699"),
}


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _title(ws: Worksheet, text: str, ncols: int) -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


def _write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: list[Sequence[Any]],
    *,
    money_cols: set[int] = frozenset(),
    pct_cols: set[int] = frozenset(),
    int_cols: set[int] = frozenset(),
    start_row: int = 2,
) -> int:
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
    ws.row_dimensions[start_row].height = 26

    r = start_row + 1
    for row_vals in rows:
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = BORDER_THIN
            col0 = c_idx - 1
            if col0 in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col0 in pct_cols:
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col0 in int_cols:
                cell.number_format = INT_FMT
                cell.alignment = Alignment(horizontal="right")
        r += 1

    last = r - 1
    for c_idx, h in enumerate(headers, 1):
        values = [row_vals[c_idx - 1] for row_vals in rows[:200]] if rows else []
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho_columna(values, h)
    ws.freeze_panes = f"A{start_row + 1}"
    if last >= start_row + 1:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last}"
    return last


# ──────────────────────────────────────────────────────────────────────────
# 🎯 MEMO EJECUTIVO — primera hoja del Excel (para el gerente).
# ──────────────────────────────────────────────────────────────────────────

def _num(v: Any) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _venta_perdida_diaria(row: dict) -> float:
    """Estima venta perdida por día para un SKU en quiebre.

    Modelo simple: el SKU venía vendiendo 'Vendido SKU S/' soles en 90 días, así
    que perder un día sin stock cuesta aproximadamente ese monto / 90. Si no hay
    monto (NULL en variant_costs/precio), usa Proyección 30d × ticket promedio
    de su categoría — pero como segundo argumento no lo tenemos acá, caemos a 0.

    El gerente lee esto como cota inferior de costo de oportunidad diario.
    """
    monto_90d = _num(row.get("Vendido SKU S/"))
    if monto_90d > 0:
        return monto_90d / 90.0
    return 0.0


def _es_urgente(row: dict) -> bool:
    """Mismo criterio que la QuiebreCard del Tablero: BESTSELLER + OPORTUNIDAD."""
    label = str(row.get("Clasificación") or "").upper()
    return "BESTSELLER" in label or "OPORTUNIDAD" in label


def _sheet_memo_ejecutivo(
    wb: Workbook,
    *,
    filtered_rows: list[dict],
    sucursal: str | None,
    brand_name: str,
    skus_con_similar: int = 0,
) -> Worksheet:
    """Crea la pestaña 🎯 Memo Ejecutivo con KPIs, top 15 y breakdowns."""
    ws = wb.create_sheet("🎯 Memo Ejecutivo")
    ws.sheet_properties.tabColor = C_TITLE_BG

    # ── 1) Banner principal (filas 1-3) ──
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 8

    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = "🎯  SKUs en Quiebre — Resumen Ejecutivo"
    cell.font = Font(name="Calibri", size=20, bold=True, color=C_TITLE_FG)
    cell.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("A3:H3")
    sub = ws["A3"]
    fecha = datetime.now().strftime("%d %b %Y").lower()
    suc_txt = f"sucursal {sucursal}" if sucursal else "todas las tiendas"
    sub.value = f"{brand_name.upper()}  ·  {suc_txt}  ·  generado el {fecha}"
    sub.font = Font(name="Calibri", size=11, bold=False, color=C_META_FG)
    sub.fill = PatternFill("solid", fgColor=C_META_BG)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── 2) KPI cards (filas 5-7) ──
    total_skus = len(filtered_rows)
    urgentes = sum(1 for r in filtered_rows if _es_urgente(r))
    vp_diaria_total = sum(_venta_perdida_diaria(r) for r in filtered_rows)
    deptos_afectados = len({str(r.get("Departamento") or "—") for r in filtered_rows})

    kpis = [
        ("SKUs EN QUIEBRE", f"{total_skus:,}".replace(",", "."), "stock = 0 con demanda"),
        ("URGENTES",        f"{urgentes:,}".replace(",", "."),   "Bestsellers + Oportunidad Perdida"),
        ("VENTA PERDIDA / DÍA", f"S/ {vp_diaria_total:,.0f}".replace(",", "."), "cota inferior del costo diario"),
        ("DEPARTAMENTOS",   f"{deptos_afectados}",               "con al menos 1 SKU en quiebre"),
    ]

    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 38
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 8

    # columnas A..H (8) → 4 KPIs en bloques de 2 columnas cada uno
    for i, (lbl, val, sub_lbl) in enumerate(kpis):
        col_start = 1 + i * 2  # A, C, E, G
        col_end = col_start + 1

        # Etiqueta arriba
        ws.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_end)
        cell = ws.cell(row=6, column=col_start)
        cell.value = lbl
        cell.font = Font(name="Calibri", size=9, bold=True, color=C_TOTAL_LBL)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Valor grande en el centro
        ws.merge_cells(start_row=7, start_column=col_start, end_row=7, end_column=col_end)
        cell = ws.cell(row=7, column=col_start)
        cell.value = val
        cell.font = Font(name="Calibri", size=24, bold=True, color=C_TOTAL_FG)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Subetiqueta debajo
        ws.merge_cells(start_row=8, start_column=col_start, end_row=8, end_column=col_end)
        cell = ws.cell(row=8, column=col_start)
        cell.value = sub_lbl
        cell.font = Font(name="Calibri", size=9, italic=True, color=C_NOTE_FG)
        cell.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Pequeño gutter entre KPI cards (col B, D, F en realidad coinciden con merge — no separan)
    # Si quisiera espacio real, tendría que usar cols A, C, E, G y dejar B, D, F en blanco angosto.
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 18

    # ── 2b) Aviso de similares (fila 10, gap antes del Top 15) ──
    # Advertencia, no exclusión: el gerente decide si compra o usa el similar.
    if skus_con_similar > 0:
        ws.merge_cells("A10:H10")
        cell = ws["A10"]
        cell.value = (
            f"⚠ {skus_con_similar} SKUs a comprar ya tienen un producto similar con stock en tienda "
            "— ver columna '⚠ Similar en tienda' en las pestañas por departamento"
        )
        cell.font = Font(name="Calibri", size=10, bold=True, color="78350F")
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[10].height = 20

    # ── 3) Top 15 SKUs por venta perdida estimada (filas 11+) ──
    ws.merge_cells("A11:H11")
    cell = ws["A11"]
    cell.value = "🚨 TOP 15 — SKUs prioritarios (por venta perdida estimada)"
    cell.font = Font(name="Calibri", size=13, bold=True, color=C_DEPT_FG)
    cell.fill = PatternFill("solid", fgColor=C_DEPT_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[11].height = 26

    # Ordenar por venta perdida diaria desc.
    top15 = sorted(filtered_rows, key=_venta_perdida_diaria, reverse=True)[:15]

    headers = ["#", "Producto", "SKU", "Sucursal", "Categoría", "Clasificación", "Proy 30d", "S/ perdido/día"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=12, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_CAT_FG)
        cell.fill = PatternFill("solid", fgColor=C_CAT_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=C_DEPT_BG))
    ws.row_dimensions[12].height = 22

    row_idx = 13
    for i, r in enumerate(top15, 1):
        producto = str(r.get("Producto") or "—")
        sku = str(r.get("Código SKU") or "")
        suc = str(r.get("Sucursal") or "—")
        cat = str(r.get("Categoría") or "—")
        clasif_full = str(r.get("Clasificación") or "—")
        clasif = clasif_full.split(":")[0].split("—")[0].strip()  # solo el chip (antes de los ":")
        proy_30 = _num(r.get("Proyección 30d"))
        vp = _venta_perdida_diaria(r)

        vals = [i, producto, sku, suc, cat, clasif, proy_30, vp]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left" if c_idx in (2, 5, 6) else "center", vertical="center")
            cell.font = Font(name="Calibri", size=10, color="1F2937")
            # Bandas zebra
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
            if c_idx == 1:  # ranking
                cell.font = Font(name="Calibri", size=10, bold=True, color=C_DEPT_BG)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == 7:  # proy 30d
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if c_idx == 8:  # S/ perdido
                cell.number_format = '"S/ "#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name="Calibri", size=10, bold=True, color=C_TOTAL_FG)
        row_idx += 1

    if not top15:
        ws.cell(row=13, column=1, value="✓ No hay SKUs en quiebre — todo el catálogo crítico tiene stock.")
        ws.cell(row=13, column=1).font = Font(italic=True, color="065F46")
        ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=8)
        row_idx = 14

    # Anchos para la tabla
    widths = {"A": 5, "B": 36, "C": 18, "D": 16, "E": 22, "F": 26, "G": 11, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── 4) Breakdown por sucursal ──
    row_idx += 2  # espacio
    base_sucursal = row_idx
    ws.merge_cells(start_row=base_sucursal, start_column=1, end_row=base_sucursal, end_column=4)
    cell = ws.cell(row=base_sucursal, column=1)
    cell.value = "🏬 Por sucursal"
    cell.font = Font(name="Calibri", size=12, bold=True, color=C_CAT_FG)
    cell.fill = PatternFill("solid", fgColor=C_SUBCAT_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[base_sucursal].height = 22

    # Header
    sh = ["Sucursal", "Quiebres", "Urgentes", "S/ perdido/día"]
    for c_idx, h in enumerate(sh, 1):
        cell = ws.cell(row=base_sucursal + 1, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_CAT_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=C_DEPT_BG))

    suc_counter: Counter[str] = Counter()
    suc_urgentes: Counter[str] = Counter()
    suc_vp: defaultdict[str, float] = defaultdict(float)
    for r in filtered_rows:
        s = str(r.get("Sucursal") or "—")
        suc_counter[s] += 1
        if _es_urgente(r):
            suc_urgentes[s] += 1
        suc_vp[s] += _venta_perdida_diaria(r)

    suc_sorted = sorted(suc_counter.items(), key=lambda x: -x[1])
    for i, (s, n) in enumerate(suc_sorted, 1):
        ri = base_sucursal + 1 + i
        ws.cell(row=ri, column=1, value=s).alignment = Alignment(horizontal="left")
        ws.cell(row=ri, column=2, value=n).number_format = "#,##0"
        ws.cell(row=ri, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=ri, column=3, value=suc_urgentes[s]).number_format = "#,##0"
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=ri, column=4, value=suc_vp[s]).number_format = '"S/ "#,##0'
        ws.cell(row=ri, column=4).alignment = Alignment(horizontal="right")
        for c in range(1, 5):
            ws.cell(row=ri, column=c).border = Border(bottom=Side(style="thin", color="E5E7EB"))
    row_idx = base_sucursal + 2 + len(suc_sorted) + 2

    # ── 5) Top departamentos afectados (mismo bloque, al lado o abajo) ──
    base_dept = row_idx
    ws.merge_cells(start_row=base_dept, start_column=1, end_row=base_dept, end_column=4)
    cell = ws.cell(row=base_dept, column=1)
    cell.value = "📂 Top 10 departamentos con quiebres"
    cell.font = Font(name="Calibri", size=12, bold=True, color=C_CAT_FG)
    cell.fill = PatternFill("solid", fgColor=C_SUBCAT_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[base_dept].height = 22

    dh = ["Departamento", "Quiebres", "Urgentes", "S/ perdido/día"]
    for c_idx, h in enumerate(dh, 1):
        cell = ws.cell(row=base_dept + 1, column=c_idx, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=C_CAT_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=C_DEPT_BG))

    dept_counter: Counter[str] = Counter()
    dept_urg: Counter[str] = Counter()
    dept_vp: defaultdict[str, float] = defaultdict(float)
    for r in filtered_rows:
        d = str(r.get("Departamento") or "—")
        dept_counter[d] += 1
        if _es_urgente(r):
            dept_urg[d] += 1
        dept_vp[d] += _venta_perdida_diaria(r)

    for i, (d, n) in enumerate(sorted(dept_counter.items(), key=lambda x: -x[1])[:10], 1):
        ri = base_dept + 1 + i
        ws.cell(row=ri, column=1, value=d).alignment = Alignment(horizontal="left")
        ws.cell(row=ri, column=2, value=n).number_format = "#,##0"
        ws.cell(row=ri, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=ri, column=3, value=dept_urg[d]).number_format = "#,##0"
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=ri, column=4, value=dept_vp[d]).number_format = '"S/ "#,##0'
        ws.cell(row=ri, column=4).alignment = Alignment(horizontal="right")
        for c in range(1, 5):
            ws.cell(row=ri, column=c).border = Border(bottom=Side(style="thin", color="E5E7EB"))
    row_idx = base_dept + 2 + min(len(dept_counter), 10) + 2

    # ── 6) Footer / call-to-action ──
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    cell = ws.cell(row=row_idx, column=1)
    cell.value = "→ Detalle por departamento en las siguientes pestañas · Venta por categoría al final"
    cell.font = Font(name="Calibri", size=10, italic=True, color=C_NOTE_FG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = 22

    # Vista limpia
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    ws.freeze_panes = "A12"  # congelar banner + KPIs

    # Imprimir en una página
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    return ws


# ──────────────────────────────────────────────────────────────────────────
# Resto del builder (legacy + jerárquico)
# ──────────────────────────────────────────────────────────────────────────

def _sheet_quiebre(wb: Workbook, skus: list[dict]) -> None:
    """SKUs que generan venta perdida HOY.

    Solo se reciben los ya filtrados a severidad 🔴 Crítico y 🟠 Alta — el
    endpoint hace el filtro contra la matriz 04 para no traer ruido.
    """
    ws = wb.create_sheet("SKUs en quiebre")
    ws.sheet_properties.tabColor = TAB_QUIEBRE
    _title(ws, "SKUs en quiebre — venta perdida HOY (críticos + altas)", 12)

    rows = [
        (
            s.get("sku"),
            s.get("producto"),
            s.get("sucursal"),
            s.get("categoria"),
            _num(s.get("stock")),
            _num(s.get("stock_almacen")),
            s.get("clasificacion"),
            s.get("severidad"),
            s.get("accion"),
            s.get("causal"),
            _num(s.get("proy_30d")),
            _num(s.get("vendido_sl")),
        )
        for s in skus
    ]
    last = _write_table(
        ws,
        [
            "SKU", "Producto", "Sucursal", "Categoría",
            "Stock", "Stock Alm.", "Clasificación original",
            "Severidad", "Acción", "Causal",
            "Proy. 30d", "Vendido S/",
        ],
        rows,
        int_cols={4, 5, 10},
        money_cols={11},
    )
    # Colorear severidad
    for r in range(3, last + 1):
        cell = ws.cell(row=r, column=8)
        v = cell.value
        if v and v in _SEV_FILLS:
            cell.fill = _SEV_FILLS[v]
            cell.font = Font(bold=True)


def _sheet_categoria(wb: Workbook, cats: list[dict]) -> None:
    """Venta por categoría con ticket promedio."""
    ws = wb.create_sheet("Venta por categoría")
    ws.sheet_properties.tabColor = TAB_CAT
    _title(ws, "Venta por categoría — ranking + ticket promedio", 6)
    rows = []
    for c in cats:
        ventas = _num(c.get("ventas")) or 0
        tickets = _num(c.get("tickets")) or 0
        ticket_prom = (ventas / tickets) if tickets > 0 else None
        rows.append((
            c.get("departamento"),
            c.get("categoria"),
            ventas,
            _num(c.get("participacion_pct")),
            tickets,
            ticket_prom,
        ))
    _write_table(
        ws,
        ["Departamento", "Categoría", "Ventas (S/)", "% participación", "Tickets", "Ticket prom. cat. (S/)"],
        rows,
        money_cols={2, 5},
        pct_cols={3},
        int_cols={4},
    )


def build_compras_catalogo_workbook(
    *,
    skus_quiebre: list[dict],
    venta_categoria: list[dict],
) -> Workbook:
    """[Legacy] Workbook plano con 2 pestañas (lista plana de quiebres + categorías).

    Mantenido por compat. La versión actual del endpoint usa
    `build_compras_catalogo_workbook_jerarquico` que aplica la estructura
    ejecutiva (1 hoja por departamento) del módulo 04b.
    """
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_quiebre(wb, skus_quiebre)
    _sheet_categoria(wb, venta_categoria)
    return wb


def build_compras_catalogo_workbook_jerarquico(
    *,
    cols: list[str],
    filtered_rows: list[dict],
    venta_categoria: list[dict],
    titulo: str,
    descripcion: str,
    sucursal_filtro: str | None,
    brand_name: str,
    similares_map: dict[str, dict] | None = None,
    show_margin: bool = False,
    extra_stock_cols: list[str] | None = None,
) -> Workbook:
    """Workbook ejecutivo (resumen + 1 hoja por Departamento) + pestaña Venta por categoría.

    Reutiliza el mismo builder que el Excel de /ventas-jerarquicas
    (`analytics.excel_executive.build_executive_workbook`) para que la jerarquía
    y los estilos coincidan. Las filas se reciben ya filtradas a quiebre real.

    show_margin      → agrega "Utilidad (S/)" y "Margen %".
    extra_stock_cols → nombres de columnas de stock de OTRAS sucursales
                       (ej. ["Stock KAWII ASAMBLEA"]). Ambas van al FINAL de la
                       tabla para no desplazar las constantes COL_* (evita el
                       desalineado de las filas de subtotal).
    """
    from analytics.excel_executive import build_executive_workbook

    # Convertir filas dict → tuplas (orden de columnas).
    rows_tuples = [tuple(row.get(c) for c in cols) for row in filtered_rows]

    wb = build_executive_workbook(
        cols=cols,
        rows=rows_tuples,
        modulo_id="compras",
        titulo=titulo,
        sql_file="(matriz 04b filtrada a quiebres reales)",
        descripcion=descripcion,
        classification_col="Clasificación",
        elapsed_seconds=0.0,
        brand_name=brand_name,
        sucursal=sucursal_filtro,
        accion_label="Compras & Catálogo",
        periodo_dias=90,
        similares_map=similares_map,
        show_margin=show_margin,
        extra_stock_cols=extra_stock_cols,
    )

    # Agregar pestaña final "Venta por categoría" con la tabla simple + ticket prom.
    _sheet_categoria(wb, venta_categoria)

    # 🎯 Memo Ejecutivo — PRIMERA hoja para el gerente (KPIs hero, top 15,
    # breakdowns por sucursal y depto). Se crea al final y se reordena con
    # `wb._sheets` para quedar como tab inicial.
    memo = _sheet_memo_ejecutivo(
        wb,
        filtered_rows=filtered_rows,
        sucursal=sucursal_filtro,
        brand_name=brand_name,
        skus_con_similar=len(similares_map or {}),
    )
    # Reordenar: memo PRIMERO, después el resto en su orden actual.
    wb._sheets = [memo] + [s for s in wb._sheets if s is not memo]
    wb.active = 0  # abrir el Excel en el memo
    return wb
